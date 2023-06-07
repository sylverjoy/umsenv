import json
import xlsxwriter
import io
import openpyxl
from django.contrib.auth.models import Group
import requests
from django.contrib import messages
from django.core.files.storage import FileSystemStorage
from django.http import HttpResponse, JsonResponse
from django.shortcuts import (HttpResponse, HttpResponseRedirect,
                              get_object_or_404, redirect, render)
from django.templatetags.static import static
from django.urls import reverse
from django.views.decorators.csrf import csrf_exempt
from django.views.generic import UpdateView
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth.decorators import login_required

from .models import *
from django.core import management
from django.core.management.commands import loaddata

#from .EmailBackend import EmailBackend
from .models import *
from .forms import *
from django.contrib import messages

from django.contrib.auth import authenticate, login, logout
from .decorators import allowed_users, unauthenticated_user,allowed_users_home
from django.http import HttpResponse
from django.views.generic import View

from django.http import FileResponse
import io
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch
from reportlab.lib.pagesizes import letter
from django.db import connections
import os
from django.db.models import Avg

from django.contrib.auth.models import User

 
def cal_cg(marks):
    if marks>=16:
        return 4.0
    elif marks>=14:
        return 3.5
    elif marks>=12:
        return 3.00
    elif marks>=11:
        return 2.50
    elif marks>=10:
        return 2.00
    elif marks>=9:
        return 1.50
    elif marks>=8:
        return 1.00
    else:
        return 0.00
    
def cal_grade(mark):
    if mark>=15:
        return "Distinction"
    elif mark>=14:
        return "Upper Credit"
    elif mark>=12:
        return "Lower Credit"
    elif mark>=10:
        return "Pass"
    else:
        return "Fail"
    
def cal_class(mark):
    if mark>=3.6:
        return "First Class"
    elif mark>=3.0:
        return "Second Class - UD"
    elif mark>=2.5:
        return "Second Class - LD"
    elif mark>=2.25:
        return "Third Class"
    elif mark>=2.0:
        return "Pass"
    else:
        return "Fail"

def cal_cgname(cg):
    if cg == 4.00:
        return "A"
    elif cg == 3.50:
        return "B+"
    elif cg == 3.00:
        return "B"
    elif cg == 2.50:
        return "C+"
    elif cg == 2.00:
        return "C"
    elif cg == 1.50:
        return "D+"
    elif cg == 1.00:
        return "D"
    elif cg == 0.00:
        return "F"


##-----------------------------LOG IN AND REGISTER--------------------------------------------------------###

@unauthenticated_user
def loginPage(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request,user)
            return redirect('home')
        else:
            messages.info(request, "Username Or Password is not Correct")
        

    return render(request, 'login_template/login1.html')

@login_required(login_url = 'login')
@allowed_users(allowed_roles=['admin'])
def registerPageTeacher(request):
    user_form = CreateUserForm(request.POST or None) 
    teacher_form = TeacherForm(request.POST or None, request.FILES or None)
    context = {'teacher_form': teacher_form,'user_form':user_form, 'page_title':'add student'}
    if request.method == 'POST':
        if user_form.is_valid and teacher_form.is_valid():
            user = user_form.save()
            teacher = teacher_form.save()
            teacher.user =user
            teacher.save()
            
            group = Group.objects.get(name = 'teacher')
            user.groups.add(group)
            messages.success(request, "Successfully Added Teacher")
        else:
            messages.success(request, "Teacher Couldn't Be Added")
            
            
    return render(request, 'registration_template/add_teacher.html',context)

@login_required(login_url = 'login')
@allowed_users(allowed_roles=['admin'])
def registerPage(request):
    user_form = CreateUserForm(request.POST or None) 
    admin_form = AdminForm(request.POST or None, request.FILES or None)
    context = {'admin_form': admin_form,'user_form':user_form, 'page_title':'add student'}
    if request.method == 'POST':
        if user_form.is_valid() and admin_form.is_valid():
            user = user_form.save()
            admin = admin_form.save()
            admin.user =user
            admin.save()
            
            group = Group.objects.get(name = 'admin')
            user.groups.add(group)
            messages.success(request, "Successfully Admin Added")
        else:
            messages.success(request, "Admin Couldn't Be Added")
    return render(request, 'registration_template/add_admin.html',context)

def logoutPage(request):
    logout(request)
    return redirect('login') 

##-----------------------------------------------LOG IN END -------------------------------------------------------###


##----------------------------------------------- HOME PAGE ----------------------------------------------------------##
@login_required(login_url = 'login')
@allowed_users_home(allowed_roles=['admin', 'teacher'])
def home(request):
    name = str(request.user.adminuser.name)
    stu_cnt = Student.objects.all().count()
    teacher_cnt =Teacher.objects.all().count()
    school_cnt = School.objects.all().count()
    admin_cnt = AdminUser.objects.all().count()
    sub_cnt = Subject.objects.all().count()
    overall_rate =Rating.objects.aggregate(Avg('rating'))
    hi = Result.objects.raw ('''
    SELECT 1 as id,dept, AVG(total) as avg from main_result group by dept ORDER BY avg DESC  LIMIT 1;
    
    ''')
    low = Result.objects.raw ('''
    SELECT 1 as id,dept, AVG(total) as avg from main_result group by dept ORDER BY avg LIMIT 1;
    
    ''')
    hi_dept =""
    low_dept =""
    hi_mark =0
    low_mark =0
    for i in hi:
        hi_dept = i.dept
        hi_mark = round((i.avg)/4.0, 2)
    for i in low:
        low_dept = i.dept
        low_mark = round((i.avg)/4.0, 2)

    
    
    print(overall_rate)
    
    if overall_rate['rating__avg'] is None:
        overall_rate = {'rating__avg': 0 }
    

    context = {'name':name,
                'stu':stu_cnt,
                'teacher':teacher_cnt,
                'school': school_cnt,
                'admin_cnt': admin_cnt,
                'sub_cnt':sub_cnt,
                'overall_rate': round(overall_rate['rating__avg'],2),
                'hi_dept': hi_dept,
                'hi_mark':hi_mark,
                'low_dept': low_dept,
                'low_mark':low_mark,                

    }
    return render(request,'admin_template/index.html',context)

@login_required(login_url = 'login')
@allowed_users(allowed_roles=['student'])
def studentHome(request):
    name = request.user.student.name
    regi = request.user.student.registration_number
    dept = request.user.student.dept
    res = RegisterTable.objects.filter(student_id = regi)
    print(res)
    Subject.objects.raw('''
        SELECT 1 as id, SUM(credit)
        FROM public.main_student JOIN public.main_result ON
        main_student.registration_number = main_result.student_id
        JOIN public.main_subject ON main_result.course_code = main_subject.course_code
        where main_student.registration_number=%s;''',[regi])

    if res == None:
        data = []
        credits = 0
        credits_passed = 0
        percent_passed_credit = 0
        percent_registered = 0
        remain_credit = 60
        remain_credit_percent= 100
        current_cgpa =0
        degree_status = "INCOMPLETE"
        remain = "Need to pass "+ str(remain_credit) +" more credits to get a degree" 
    
    else:
        credits = 0
        for r in res:
            sub = Subject.objects.filter(course_code = r.subject.course_code).first()
            print(sub)
            print(sub.credit)
            credits += sub.credit
        print(credits)
        #credits = Subjects.objects.raw('''
        #SELECT 1 as course_code, SUM(credit)
        #FROM public.main_student JOIN public.main_result ON
        #main_student.registration_number = main_result.student_id
        #JOIN public.main_subject ON main_result.course_code = main_subject.course_code
        #where main_student.registration_number=%s;''',[regi])[0].sum

        credits_passed = 0
        res = Result.objects.filter(student = regi).all()

        for r in res:
            if r.total >= 10:
                credits_passed+= r.course_code.credit

        degree_status = ""


        upper =0
        lower = 0
        for r in res:
            upper =upper+ r.total
            lower += 1

        if lower == 0:
            lower = 1
        
        if credits_passed is None:
            credits_passed = 0

        current_cgpa = round(upper/lower,2)
        percent_passed_credit = ((credits_passed)*100)/60


        percent_registered = ((credits)*100)/60

        remain_credit = 60 - credits_passed

        remain_credit_percent = ((remain_credit)*100)/60
        remain =""
        if(credits_passed<60):
            degree_status = "INCOMPLETE"
            remain = "Need to pass "+ str(remain_credit) +" more credits to get a degreee" 
        else:
            degree_status = "COMPLETE"
            remain = "Congratulations!!! You are a YIBS "+ str(dept) +" Graduate"  

    context ={'name':name, 
    'credits': credits,
    'percent_registered': percent_registered,
    'credits_passed' : credits_passed,
    'percent_passed_credit': percent_passed_credit,
    'current_cgpa': current_cgpa,
    'degree_status': degree_status,
    'remain': remain,
    
    
    }
    return render(request,'student_template/index.html',context)

@login_required(login_url = 'login')
def teacher_home(request):
    name = str(request.user.teacher.name)
    dept = str(request.user.teacher.dept_id)
    t_id = str(request.user.teacher.teacher_id)
    num_of_regi_sub = AssignedTeacher2.objects.filter(teacher_id = t_id).count()
    context ={ 'name':name,
                'dept':dept,
                't_id': t_id,
                'n_ass': num_of_regi_sub,

    }
    return render(request,'teacher_template/index.html',context)

@login_required(login_url = 'login')
def courseWiseParticipation(request, *args, **kwargs):
    t_id = str(request.user.teacher.teacher_id)
    attendance = Result.objects.raw('''
    SELECT 1 as id, main_assignedteacher2.course_code as sn , SUM(attendence) as attend FROM
    public.main_assignedteacher2 JOIN public.main_result ON
    main_assignedteacher2.course_code = main_result.course_code AND student_dept = dept
	where main_assignedteacher2.teacher_id=%s
	group by main_assignedteacher2.course_code;''',[t_id])
    data =[]
    labels =[]
    for i in attendance:
        labels.append(i.sn)
        data.append(i.attend)

    return JsonResponse(data={
        'labels': labels,
        'data':data,
    })



@login_required(login_url = 'login')
def course_wise_performance(request, *args, **kwargs):
    t_id = str(request.user.teacher.teacher_id)
    attendance = Result.objects.raw('''
    SELECT 1 as id, main_assignedteacher2.course_code as sn , SUM(total) as total, COUNT(*) as cnt FROM
    public.main_assignedteacher2 JOIN public.main_result ON
    main_assignedteacher2.course_code = main_result.course_code AND student_dept = dept
	where main_assignedteacher2.teacher_id=%s
	group by main_assignedteacher2.course_code;''',[t_id])
    data =[]
    labels =[]

    for i in attendance:
        labels.append(i.sn)
        data.append((i.total/(i.cnt*100.0))*10)

    return JsonResponse(data={
        'labels': labels,
        'data':data,
    })

@login_required(login_url = 'login')
def session_wise_courses(request, *args, **kwargs):
    t_id = str(request.user.teacher.teacher_id)
    attendance = Result.objects.raw('''
    SELECT 1 as id, semester as sn , COUNT(*) as cnt FROM
    public.main_assignedteacher2
    JOIN main_subject ON main_assignedteacher2.course_code = main_subject.course_code
	where main_assignedteacher2.teacher_id=%s
	group by semester;''',[t_id])
    data =[]
    labels =[]

    for i in attendance:
        labels.append(i.sn)
        data.append(i.cnt)

    return JsonResponse(data={
        'labels': labels,
        'data':data,
    })

#--------------------------------------------*****GET*****-------------------------------------------------------------#

@login_required(login_url = 'login')
def get_att(request, *args, **kwargs):
    regi = request.user.student.registration_number
    attendance = Result.objects.raw('''
    SELECT 1 as id, subject_name as sn , attendence as attend FROM
    public.main_student JOIN public.main_result ON
    main_student.registration_number = main_result.student_id
    JOIN public.main_subject ON main_result.course_code = main_subject.course_code
    where main_student.registration_number=%s;''',[regi])
    data =[]
    labels =[]
    for i in attendance:
        labels.append(i.sn)
        data.append(i.attend)

    return JsonResponse(data={
        'labels': labels,
        'data':data,
    })
@login_required(login_url = 'login')
def get_subtype(request, *args, **kwargs):
    regi = request.user.student.registration_number
    subtype = Result.objects.raw('''
    SELECT 1 as id, subtype, SUM(total) as sum_marks, count(subtype) as cnt FROM
    public.main_student JOIN public.main_result ON
    main_student.registration_number = main_result.student_id
    JOIN public.main_subject ON main_result.course_code = main_subject.course_code
	where registration_number = %s
    group by subtype;''',[regi])

    data =[]
    labels =[]
    for i in subtype:
        labels.append(i.subtype)
        cntt = min(i.cnt*.75,3)
        print(i.cnt)
        print(i.sum_marks)
        data.append((i.sum_marks)*8/(i.cnt*100) + cntt)

    return JsonResponse(data={
        'labels': labels,
        'data':data,
    })
def getting_json(subtype, regi):

    marksObj = Result.objects.raw('''
    SELECT 1 as id, subject_name, main_subject.course_code as cc, total FROM
    public.main_student JOIN public.main_result ON
    main_student.registration_number = main_result.student_id
    JOIN public.main_subject ON main_result.course_code = main_subject.course_code
	where registration_number = %s and subtype = %s;''',[regi,subtype])
    

    subject_name =[]
    course_code =[]
    marks = []
    attr = []
    attr.append("subject_name")
    attr.append("course_code")
    attr.append("marks")
    json_res =[]
    for i in marksObj:
        obj = {}
        obj[attr[0]] = i.subject_name
        print(i.subject_name)
        obj[attr[1]] = i.cc
        
        cg = cal_cg(i.total)
        print(i.total)
        obj[attr[2]]= cg
        json_res.append(obj) 
    return json_res


def getting_json_result(regi):

    marksObj = Result.objects.filter(student = regi)

    c_ss = SemesterSession.objects.filter(active = 'Yes').first()
    print(c_ss.ss_id)
    
    pub_marks = []

    if c_ss.results_published == "No":
        print('exams not published')
        for m in marksObj:
            print(m.sem_ses)
            if str(m.sem_ses) != str(c_ss.ss_id):
                print(m)
                pub_marks.append(m)
    else:
        pub_marks = marksObj
    attr = []
    
    attr.append("subject_name")
    attr.append("course_code")
    attr.append("credit")
    attr.append("theory")
    attr.append("tt")
    attr.append("total")
    #attr.append("cgpa")
    json_res =[]
    for i in pub_marks:
        obj = {}
        obj[attr[0]] = i.course_code.subject_name
        obj[attr[1]] = i.course_code.course_code
        obj[attr[2]]= i.course_code.credit
        obj[attr[3]] = i.theory_marks
        obj[attr[4]] = i.term_test
        obj[attr[5]] = i.total

        #cgpa = cal_grade(i.total)
        #obj[attr[6]]= cgpa
        json_res.append(obj) 
        
    return json_res

def getting_students_json():
    Studs = Student.objects.all().order_by('dept', 'name')

    attr = [] 
    attr.append("Matricule")
    attr.append("Name")
    attr.append('Department')
    attr.append("Degree Pursued")
    attr.append("Level")

    json_res = []
    for stud in Studs:
        obj = {}
        obj[attr[0]] = stud.registration_number
        obj[attr[1]] = stud.name
        obj[attr[2]] = stud.dept.dept_id
        obj[attr[3]] = stud.degree_pursued.deg_id
        obj[attr[4]] = stud.level
        json_res.append(obj)
    
    return json_res

def teacher_subject_list(request):


    t_id  = request.user.teacher.teacher_id

    data = AssignedTeacher2.objects.filter(teacher = t_id)

    attr = []
    attr.append("name")
    attr.append("cc")
    attr.append("subtype")
    json_res =[]
    for i in data:
        obj = {}
        obj[attr[0]] = i.course.subject_name
        obj[attr[1]] = i.course.course_code
        obj[attr[2]] = i.course.subtype
        
        
        json_res.append(obj) 
        
    return JsonResponse(json_res, safe = False) 

@login_required(login_url = 'login')
def get_subtype_networking_marks(request, *args, **kwargs):
    regi = request.user.student.registration_number
    subtype = "Networking"
    
    json_res = getting_json(subtype, regi)
        

    return JsonResponse(json_res, safe = False)  

@login_required(login_url = 'login')
def get_subtype_dbms_marks(request, *args, **kwargs):
    regi = request.user.student.registration_number
    subtype = "DBMS"
    json_res = getting_json(subtype, regi)
    return JsonResponse(json_res, safe = False) 

@login_required(login_url = 'login')
def get_subtype_ai_marks(request, *args, **kwargs):
    regi = request.user.student.registration_number
    subtype = "AI"
    json_res = getting_json(subtype, regi)
    return JsonResponse(json_res, safe = False)


@login_required(login_url = 'login')
def get_subtype_programming_marks(request, *args, **kwargs):
    regi = request.user.student.registration_number
    subtype = "Programming"
    json_res = getting_json(subtype, regi)
    return JsonResponse(json_res, safe = False)

@login_required(login_url = 'login')
def get_subtype_sys_n_media_marks(request, *args, **kwargs):
    regi = request.user.student.registration_number
    subtype = "System & Multimedia"
    json_res = getting_json(subtype, regi)
    return JsonResponse(json_res, safe = False) 


@login_required(login_url = 'login')
def get_subtype_project_marks(request, *args, **kwargs):
    regi = request.user.student.registration_number
    subtype = "Project"
    json_res = getting_json(subtype, regi)
    return JsonResponse(json_res, safe = False) 


@login_required(login_url = 'login')
def get_all_the_marks(request, *args, **kwargs):
    regi = request.user.student.registration_number
    json_res = getting_json_result(regi)
    return JsonResponse(json_res, safe = False)

@login_required(login_url = 'login')
def get_all_the_students(request, *args, **kwargs):
    json_res = getting_students_json()
    return JsonResponse(json_res, safe = False)

def see_registration_status(request, *args, **kwargs):
    regi = str(request.user.student.registration_number)
    dep = str(request.user.student.dept)
    register = Result.objects.raw('''
    SELECT 1 as id, status, subject_id, dept_id as sub, dept_id as teacher FROM main_registertable
	where student_id = %s ''',[regi])

    for i in register:
        c_id = i.subject_id
        subject_name = Subject.objects.get(course_code = c_id).subject_name  
        #assi_tea = AssignedTeacher2.objects.get(course_code = c_id, student_dept = dep).teacher_id
        #teacher_name = Teacher.objects.get(teacher_id = assi_tea).name
        #print(teacher_name)
        #i.sub = subject_name
        #i.teacher = teacher_name
    attr=[]
    attr.append("course_code")
    #attr.append("subject_name")
    attr.append("status")
    #attr.append("teacher")
    json_res =[]
    for i in register:
        obj = {}
        obj[attr[0]] = i.subject_id
        #obj[attr[1]] = i.sub
        obj[attr[1]]= i.status
        #obj[attr[3]]= i.teacher
        json_res.append(obj)
    return JsonResponse(json_res, safe= False)
    

#------------------------------------------------------***VIEW***-----------------------------------------------------------
@login_required(login_url = 'login')
def full_attendance(request):
    return render(request,'student_template/full_attendance.html')

@login_required(login_url = 'login')
def full_marksheet(request):
    return render(request,'student_template/full_marksheet.html')

@login_required(login_url = 'login')
def all_students(request):
    return render(request,'admin_template/view_studs.html')


@login_required(login_url = 'login')
def full_skillset(request):
    return render(request,'student_template/full_skillset.html')

@login_required(login_url = 'login')
def subject_ranksheet(request):
    regi = request.user.student.registration_number
        
    data = Result.objects.raw('''
    SELECT 1 as id, course_code FROM main_result
	where student_id = %s''',[regi])

    context={'course':data, 'regi': regi} 

    if request.method == 'POST':
        course_id = request.POST.get('course_code')
        marksObj = Result.objects.raw('''
        SELECT 1 as id,main_student.name, main_student.registration_number as regi, total as marks FROM
        public.main_student JOIN public.main_result ON
        main_student.registration_number = main_result.student_id
        JOIN public.main_subject ON main_result.course_code = main_subject.course_code
	    where main_subject.course_code = %s order by marks DESC;''',[course_id])
        cnt=1
        your_rank = 0
        regi1 = request.user.student.registration_number
        for i in marksObj:
            if i.regi == regi1:
                your_rank = cnt  
            i.id =cnt
            cnt = cnt+1
        subject_name = Subject.objects.get(course_code = course_id).subject_name
        context={'data':marksObj,'course_id':course_id,'subject_name':subject_name, 'rank':your_rank} 
        return render(request, 'student_template/rank_result2.html',context)


    return render(request, 'student_template/rank_result.html',context)

from django.http import HttpResponse
from django.views.generic import View

from main.utils import html_to_pdf 
from django.template.loader import render_to_string
from django.core.files import File
import os
class GeneratePdf(View):
     def get(self, request, *args, **kwargs):
        regi = request.user.student.registration_number
        name = request.user.student.name
        phone = request.user.student.phone
        email = request.user.email
        dept = request.user.student.dept.name
        data = Result.objects.filter(student = regi).all()

        upper =0
        count = 0
        lower = 0
        cgpa = 0
        status =""
        cnt = 1
        for k in data:
            if k.total>=10:
                lower =lower + k.course_code.credit
            upper+=k.total
            count+=1
            k.id = cnt
            cnt = cnt+1
        
        if lower == 0:
            cgpa =0
            
        else:
            cgpa = round(upper/count, 2)

        status = cal_grade(cgpa)

        '''if lower<60:
            status = "Incomplete"
        else:
            status = "Complete"'''

        for i in data:
            i.student_id = cal_grade(i.total)
                    

        module_dir = os.path.dirname(__file__)  # get current directory
        file_path1 = os.path.join(module_dir, 'templates/student_template/generate_result_pdf_temp.html')
        file_path2 = os.path.join(module_dir, 'templates/student_template/generate_result_pdf.html')

        pwd = os.path.dirname(__file__)
        open(file_path1, "w").write(render_to_string(file_path2, {'data': data,'regi':regi,'name':name, 'phone': phone, 'email':email,'cgpa':cgpa,'status':status,'dept':dept }))

        # Converting the HTML template into a PDF file
        pdf = html_to_pdf(file_path1)
        
        # rendering the template
        return HttpResponse(pdf, content_type='application/pdf')

def my_trans(request):
    data = AcademicYear.objects.all()

    context = {'ays': data}

    if request.method == 'POST':
        aca = request.POST.get('ay')
        return redirect(reverse('generate_trans', kwargs={"ay": aca}))

    return render(request, 'student_template/trans.html',context)

def admin_trans(request):
    data = AcademicYear.objects.all()

    depts = Dept.objects.all()

    context = {'ays': data, 'ds': depts}

    if request.method == 'POST':
        aca = request.POST.get('ay')
        dept = request.POST.get('dept')
        return redirect(reverse('generate_trans', kwargs={"ay": aca, "dept": dept}))

    return render(request, 'admin_template/trans.html',context)


class GenerateTrans(View):
    ay = None
    d = None
    studs = []
    def get(self, request, *args, **kwargs):
        self.ay = self.kwargs.get('ay', None)
        self.d = self.kwargs.get('dept', None)

        if self.d != None:
            self.studs = Student.objects.filter(dept = self.d).all()
            zip_name = "Transcripts" + "_" + self.d + "_" + self.ay + ".zip"
        else:
            self.studs.append(request.user.student)
        
        students = self.studs

        head = Image.objects.filter(title = 'trans_head').first()

        pdfs = []
        names = []

        from io import BytesIO
        from django.template.loader import get_template
        from xhtml2pdf import pisa  

        for stud in students:

            regi = stud.registration_number
            name = stud.name
            phone = stud.phone
            email = stud.user.email
            dept = stud.dept.name
            program = stud.degree_pursued
            dob = stud.dob
            pob = stud.pob
            data1 = Result.objects.filter(student = regi).all()

            ay = str(self.ay)

            f_name = "Transcript " + "_" + name + "_" + program.deg_id + "_" + ay + ".pdf"

            data = []
            for d in data1:
                if str(d.sem_ses.session) == ay:
                    data.append(d)


            if len(data) == 0:
                messages.success(request,"You have no results for the choosen Academic Year")
            else:
                sem1 = []
                gpa_sum_s1 = 0
                cred_total_s1 = 0
                cred_earned = 0
                gpa_sem1 = 0
                count = 0
                for d in data:
                    if str(d.sem_ses.semester) == 'Semester 1':
                        sem1.append(d)
                
                level = sem1[0].level
                cnt = 1
                for s in sem1:
                    count+=1
                    gpa_sum_s1+= s.total
                    cred_total_s1+= s.course_code.credit
                    if s.absent == "Yes":
                        s.student_id = "X"
                        s.gp = "X"
                        s.gd = "X"
                        s.wp = "X"
                        s.total = "X"
                    else:
                        s.student_id = cal_grade(s.total)
                        s.id = cnt
                        s.gp = cal_cg(s.total)
                        s.gd = cal_cgname(s.gp)
                        s.wp = s.course_code.credit * s.gp
                        s.total = s.total * 5 
                        if s.resited == "Yes" and program.deg_id == "BTECH":
                            s.total = str(s.total) + " * "

                    cnt = cnt+1
                    
                    if s.total != "X" and float(str(s.total).split(" ")[0]) >= 50:
                        cred_earned+= s.course_code.credit

                wpt_s1 = 0
                gpt_s1 = 0
                for s in sem1:
                    if s.wp != "X"  and s.gp != "X":
                        wpt_s1+= s.wp
                        gpt_s1+= s.gp
                
                gpa_sem1 = round(gpa_sum_s1/count, 2)
                gp_s1 = round(gpt_s1/count, 2)
                grade_sem1 = cal_grade(gpa_sem1)

                sem2 = []
                gpa_sum_s2 = 0
                cred_total_s2 = 0
                gpa_sem2 = 0
                count = 0
                for d in data:
                    if str(d.sem_ses.semester) == 'Semester 2':
                        sem2.append(d)
                cnt = 1
                for s in sem2:
                    count+=1
                    gpa_sum_s2+= s.total
                    cred_total_s2+= s.course_code.credit
                    if s.absent == "Yes":
                        s.student_id = "X"
                        s.gp = "X"
                        s.gd = "X"
                        s.wp = "X"
                        s.total = "X"
                    else:
                        s.student_id = cal_grade(s.total)
                        s.id = cnt
                        s.gp = cal_cg(s.total)
                        s.gd = cal_cgname(s.gp)
                        s.wp = s.course_code.credit * s.gp
                        s.total = s.total * 5 
                        if s.resited == "Yes" and program.deg_id == "BTECH":
                            s.total = str(s.total) + " * "

                    cnt = cnt+1

                    if s.total != "X" and float(str(s.total).split(" ")[0]) >=50:
                        cred_earned+= s.course_code.credit

                wpt_s2 = 0
                gpt_s2 = 0
                for s in sem2:
                    if s.wp != "X"  and s.gp != "X":
                        wpt_s2+= s.wp
                        gpt_s2+= s.gp
                
                gpa_sem2 = round(gpa_sum_s2/count, 2)
                gp_s2 = round(gpt_s2/count, 2)
                grade_sem2 = cal_grade(gpa_sem2)

                year_gpa = round((gpa_sem1 + gpa_sem2)/2, 2)
                year_grade = cal_grade(year_gpa)
                yr_gp = round((gp_s1 + gp_s2)/2, 2)

                cred_attempted = cred_total_s1 + cred_total_s2

                deg_class = cal_class(yr_gp)
                            

                module_dir = os.path.dirname(__file__) 
                # get current directory
                if program.deg_id == "BTECH":
                    file_path1 = os.path.join(module_dir, 'templates/student_template/my_transcript_btech_temp.html')
                    file_path2 = os.path.join(module_dir, 'templates/student_template/my_transcript_btech.html')
                    #head = Image.objects.filter(title = 'trans_head_btech').first()
                else:
                    file_path1 = os.path.join(module_dir, 'templates/student_template/my_transcript_temp.html')
                    file_path2 = os.path.join(module_dir, 'templates/student_template/my_transcript.html')

                pwd = os.path.dirname(__file__)
                open(file_path1, "w").write(render_to_string(file_path2, {'data': sem1,
                                                                            'data1': sem2,
                                                                            'regi':regi,
                                                                            'name':name,
                                                                            'phone': phone,
                                                                            'email':email,
                                                                            'cgpa':gpa_sem1,
                                                                            'status':grade_sem1,
                                                                            'gp1': gp_s1,
                                                                            'credt1': cred_total_s1,
                                                                            'wpt1':wpt_s1,
                                                                            'cgpa1':gpa_sem2,
                                                                            'status1':grade_sem2,
                                                                            'gp2': gp_s2,
                                                                            'credt2':cred_total_s2,
                                                                            'wpt2':wpt_s2,
                                                                            'dept':dept,
                                                                            'ay': ay,
                                                                            'year_gpa': year_gpa,
                                                                            'year_grade': year_grade,
                                                                            'yr_gp': yr_gp,
                                                                            'deg_class': deg_class,
                                                                            'tcred': cred_attempted,
                                                                            'tcred_e': cred_earned,
                                                                            'program':program,
                                                                            'dob':dob,
                                                                            'pob':pob,
                                                                            'head': head,
                                                                            'level': level
                                                                            }))

                # Converting the HTML template into a PDF file
                
                pdf = html_to_pdf(file_path1)

                pdfs.append(pdf)
                names.append(f_name)

        if len(pdfs) > 1:
            import zipfile
            output = io.BytesIO()
            with zipfile.ZipFile(output, 'w') as zf:
                for n,p in zip(names, pdfs):
                    zf.writestr(n,p)
                zf.close()

            response = HttpResponse(
            output.getvalue(),
                content_type="application/x-zip-compressed",
            )
            response["Content-Disposition"] = "attachment; filename=" + zip_name
                    
                    # rendering the template
            return response
        else:
            response = HttpResponse(
            pdfs[0],
                content_type="application/pdf",
            )
            response["Content-Disposition"] = "attachment; filename=" + names[0]

            return response
    


##-----------------------------------------SEARCH------------------------------------####


@login_required(login_url = 'login')
@allowed_users(allowed_roles=['teacher'])
def search_result1(request):
    #t_id = request.user.teacher.teacher_id
    data = Subject.objects.all()

    context = {  'data':data }
    
    if request.method == 'POST':
        regi= request.POST.get('course_code')
        xx = regi.split(',')
        return redirect(reverse('search_result', kwargs={"course_code": xx[0], "dept" : xx[1]}))


    return render(request,'teacher_template/search_result1.html',context)





@login_required(login_url = 'login')
@allowed_users(allowed_roles=['teacher'])
def search_student_registered(request):

    t_id = str(request.user.teacher.teacher_id)
    data = AssignedTeacher2.objects.filter(teacher_id = t_id)
    
    context = {'data':data}
    
    if request.method == 'POST':
        course_id = request.POST.get('course_code')
        xx = course_id.split(",")
        print(xx)

        return redirect(reverse('add_result', kwargs= {"dept": xx[1], "course_id": xx[0]}))
    return render(request,'teacher_template/search_student_registered.html',context)
        
###-----------------------------------SEARCH END------------------------------------------------------------------------#

@login_required(login_url = 'login')
@allowed_users(allowed_roles=['teacher'])
def search_result(request, course_code, dept):
    data = Result.objects.filter(course_code = course_code, dept = dept)

    context={'course':data, 'course_code': course_code, 'dept': dept} 
    if request.method == 'POST':
        registration_number = request.POST.get('registration_number')
        obj = Result.objects.get(student_id = registration_number ,course_code = course_code, dept = dept )
        id = int(obj.id)
        print(id)
        return redirect(reverse('update_result', kwargs= {"result_id": id,"course_code": course_code}))
        
    return render(request,'teacher_template/search_result.html',context)



@login_required(login_url = 'login')
@allowed_users(allowed_roles=['teacher'])
def update_result(request, result_id, course_code):
    result = get_object_or_404(Result, id =result_id)
    stu = result.student_id
    course_id = result.course_code
    dept = result.dept
    theory_marks = result.theory_marks
    attendence = result.attendence
    term_test = result.term_test
    total = result.total
    print(total)
    form = UpdateForm(request.POST or None, instance = result)
    regi = result.student_id

    context = {'form':form, 
               'regi': regi, 
              'course_id': course_id, 
    }
    
    if form.is_valid():
        tt = form.instance.term_test
        aa = form.instance.attendence
        theory= form.instance.theory_marks
        form.instance.total = round((theory + tt +aa)/5)
        form.save()
        messages.success(request,"Marks Edited %s Student's %s Course"%(regi,course_id))
        return redirect('home')
    
    return render (request, 'teacher_template/update_result.html',context)


#----------------------------------------------** ADD ** -------------------------------------------------------------------------------------------------------


@login_required(login_url = 'login')
@allowed_users(allowed_roles=['admin'])
def add_student(request):
    user_form = CreateUserForm(request.POST or None) 
    student_form = StudentForm(request.POST or None, request.FILES or None)
    context = {'student_form': student_form,'user_form':user_form, 'page_title':'add student'}
    if request.method == 'POST':
        if user_form.is_valid and student_form.is_valid():
            user = user_form.save()
            student = student_form.save()
            student.user =user
            student.save()
            
            group = Group.objects.get(name = 'student')
            user.groups.add(group)
            messages.success(request, "Successfully Student Added")
        else:
            messages.error(request, "Could Not Add")
    return render(request, 'student_template/add_student.html',context)


@login_required(login_url = 'login')
@allowed_users(allowed_roles=['teacher'])
def add_result(request, dept, course_id):

    register = RegisterTable.objects.filter(subject_id = course_id, dept_id = dept, status = 'Approved')
    context = {'data':register}
    if request.method == 'POST':
        regi = request.POST.get('registration_number')
        
        obj = Result.objects.filter(student_id = regi ,course_code = course_id).first()
        if obj != None:
            id = obj.id
            messages.info(request, "Result Already Exist, You Can Update That Result Here")
            return redirect(reverse('update_result', kwargs= {"result_id": id, "course_code": course_id}))
        else:
            return redirect(reverse('add_result2', kwargs= {"regi": regi, "cour_id": course_id}))

    return render(request, 'teacher_template/add_result.html',context)


@login_required(login_url = 'login')
@allowed_users(allowed_roles=['admin'])
def add_subject(request):
    subject_form = AddSubjectForm(request.POST or None, request.FILES or None)
    context = {'subject_form': subject_form, 'page_title':'add subject'}
    if request.method == 'POST':
        if subject_form.is_valid():
            subject = subject_form.save()
            subject.save()
            
            messages.success(request, "Successfully Added Subject")
        else:
            messages.error(request, "Could Not Add")

    return render(request, 'admin_template/add_subject.html',context)



@login_required(login_url = 'login')
@allowed_users(allowed_roles=['teacher'])
def add_j(request):
    t_id = str(request.user.teacher.teacher_id)
    print(t_id)
    data = AssignedTeacher2.objects.filter(teacher_id = t_id)

    context={'course':data} 
    if request.method == 'POST' and request.FILES['myfile']:
        myfile = request.FILES['myfile']
        code = request.POST.get('course_code')
        xx =code.split(",")
        print(xx)
        course_cd = xx[0]
        dept_id =xx[1]
        fs = FileSystemStorage()
        filename = fs.save(myfile.name, myfile)
        module_dir = 'C:\\Users\\neyamul\\Projects\\university_management_system_extend\\university_management_system\\media'
        file_path =os.path.join(module_dir, filename)

        f =open(file_path)
        
        y=json.load(f)


        for i in range(0, len(y)):
            
            course = course_cd
            dept = dept_id
            regi = y[i]["student_id"]
            theory_mar =y[i]["theory_marks"]
            term_tes =y[i]["term_test"]
            attendence = y[i]["attendence"]

            total_marks = round((float(theory_mar)/100.0)*70+(float(term_tes)/30.0)*20+attendence)

            register1 = RegisterTable.objects.filter(student_id = regi, dept_id = dept, subject_id = course_cd).first()

            if register1== None:
                messages.error(request," %s student did not sent  register request for %s course "% (regi, course_cd))
                continue
            
            if register1.status == 'Pending':
                messages.error(request," %s student registration for %s course is pending, Approve first "% (regi, course_cd))
                continue

            if register1.status == 'Rejected':
                messages.error(request," %s student registration for %s course is Rejected "% (regi, course_cd))
                continue

            cd = Result.objects.filter(student_id = regi, course_code = course).first()
            if cd != None:
                messages.error(request," %s student's  %s course's result already here "% (regi, course))
                continue
            
            sd = Student.objects.filter(registration_number = regi).first()
            if sd == None:
                messages.error(request," %s student is not registered in %s Department "% (regi, dept_id))
                continue
            
            sub = Result(
                course_code =course,
                theory_marks = theory_mar,
                term_test = term_tes,
                attendence = attendence,
                dept = dept_id,
                student_id = regi,
                total = total_marks,

            )
            sub.save()
            messages.success(request," %s student's  %s course's result added "% (regi, course))
            
        messages.success(request,"Successfully Added Result for %s course"% (course))  
        return redirect('home') 
            

    return render(request,'teacher_template/add_json.html',context)


##Extracting result template
@login_required(login_url = 'login')
@allowed_users(allowed_roles=['teacher'])
def extract_temp(request):
    t_id = str(request.user.teacher.teacher_id)
    print(t_id)
    data1 = AssignedTeacher2.objects.filter(teacher_id = t_id)

    c_ss = SemesterSession.objects.filter(active = 'Yes').first()
    print(c_ss)
    print(c_ss.ca_deadline)
    disabled = ""
    data = []
    for d in data1:
        if d.course.semester == c_ss.semester :
            data.append(d)

    from datetime import datetime
    if datetime.now().date() > c_ss.ca_deadline :
        print(True)

        disabled = "disabled"
        messages.success(request,"Deadline to Submit CA Results has passed. You can't download any templates anymore.")

    context={'course':data, 'disabled': disabled} 
    if request.method == 'POST':
        code = request.POST.get('course_code')
        xx =code.split(",")
        print(xx)
        course_cd = xx[0]
        dept_id =xx[1]

        matricules = []
        names = []
        fields = ["Matricule", "Names", "CA"]

        register1 = RegisterTable.objects.filter(subject_id = course_cd).all()
        for r in register1:
            matricules.append(r.student)

            n = Student.objects.filter(registration_number = r.student).first().name
            names.append(n)

        course_name = Subject.objects.filter(course_code = course_cd).first().subject_name
        wb_name = 'CA Results_' + str(course_name) + '.xlsx'

        output = io.BytesIO()
        wb = xlsxwriter.Workbook(output)
        ws = wb.add_worksheet('Sheet1')
        
        r1 = 0
        c1 = 0

        for f in fields:
            ws.write(r1, c1, f)
            c1+=1
        r2 = 1
        c2 = 0

        import itertools
        for (m, n) in zip(matricules, names):
            ws.write(r2, c2, str(m))
            ws.write(r2,c2 + 1, str(n))
            r2+=1

        wb.close()
        
        output.seek(0)

        response = HttpResponse(
            output,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = "attachment; filename=%s" % wb_name

        return response
            

    return render(request,'teacher_template/extract_temp.html',context)

@login_required(login_url = 'login')
@allowed_users(allowed_roles=['teacher'])
def extract_temp_exam(request):
    t_id = str(request.user.teacher.teacher_id)
    print(t_id)
    data1 = AssignedTeacher2.objects.filter(teacher_id = t_id)

    c_ss = SemesterSession.objects.filter(active = 'Yes').first()
    print(c_ss)

    disabled = ""
    data = []
    for d in data1:
        if d.course.semester == c_ss.semester :
            data.append(d)

    context={'course':data, 'disabled': disabled} 
    if request.method == 'POST':
        code = request.POST.get('course_code')
        xx =code.split(",")
        print(xx)
        course_cd = xx[0]
        dept_id =xx[1]

        codes = []
        fields = ["Code", "Exam Mark"]

        register1 = RegisterTable.objects.filter(subject_id = course_cd).all()
        for r in register1:
            c = ExamCode.objects.filter(student = r.student, sem_ses = c_ss, subject = r.subject).first().code
            codes.append(c)

        course_name = Subject.objects.filter(course_code = course_cd).first().subject_name
        wb_name = 'Exam Results_' + str(course_name) + '_' + str(c_ss.ss_id) + '.xlsx'

        output = io.BytesIO()
        wb = xlsxwriter.Workbook(output)
        ws = wb.add_worksheet('Sheet1')
        
        r1 = 0
        c1 = 0

        for f in fields:
            ws.write(r1, c1, f)
            c1+=1
        r2 = 1
        c2 = 0

        
        for c in codes:
            ws.write(r2, c2, str(c))
            r2+=1

        wb.close()
        
        output.seek(0)

        response = HttpResponse(
            output,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = "attachment; filename=%s" % wb_name

        return response
            

    return render(request,'teacher_template/extract_temp_exam.html',context)

@login_required(login_url = 'login')
@allowed_users(allowed_roles=['admin'])
def extract_res_stat(request):

    ss = SemesterSession.objects.all()
    depts = Dept.objects.all()
    levs = ['HND1','HND2','BTECH']


    context={'sems':ss, 'depts': depts, 'levs': levs} 
    if request.method == 'POST':
        sem = request.POST.get('sem')
        dept = request.POST.get('dept')
        lev = request.POST.get('lev')

        res = Result.objects.filter(sem_ses = sem, level = lev, dept = dept).all()

        print(res)

        fields = ['Course Code', 'Course Title', 'Credit Value', 'Course Instructor', 'Registered', 'Examined', 'Passed', 'Failed', '% Passed']

        codes = []
        titles = [] 
        creditvs = [] 
        instructors = [] 
        reg = []
        examd = []
        passed = []
        failed = []
        per_passed = []

        for r in res:
            if r.course_code.course_code in codes:
                continue
            else:
                codes.append(r.course_code.course_code)
                titles.append(r.course_code.subject_name)
                creditvs.append(r.course_code.credit)
                instructors.append(AssignedTeacher2.objects.filter(dept = dept, course = r.course_code).first().teacher)
                reg.append(RegisterTable.objects.filter(sem_ses = sem, dept = dept, subject = r.course_code).all().count())
                ex = res.filter(course_code = r.course_code).count()
                examd.append(ex)
                psd = res.filter(course_code = r.course_code, total__gte = 10).all().count()
                passed.append(psd)
                failed.append(res.filter(course_code = r.course_code, total__lt = 10).all().count())
                per_passed.append(round((psd / ex )*100,2))

        wb_name = 'Results_Stats_' + str(dept) + '_' + str(lev) + '_' + str(sem) + '.xlsx'

        output = io.BytesIO()
        wb = xlsxwriter.Workbook(output)
        ws = wb.add_worksheet('Sheet1')

        ws.write(0,0, 'Results Statistics ' + dept + ' for ' + sem + '')
        
        r1 = 1
        c1 = 0

        for f in fields:
            ws.write(r1, c1, f)
            c1+=1

        r2 = 2
        c2 = 0
        for (c,t,cv,i,r,e,p,fd,pp) in zip(codes, titles, creditvs, instructors, reg, examd, passed, failed, per_passed):
            ws.write(r2, c2, str(c))
            ws.write(r2, c2 + 1, str(t))
            ws.write(r2, c2 + 2, cv)
            ws.write(r2, c2 + 3, str(i))
            ws.write(r2, c2 + 4, r)
            ws.write(r2, c2 + 5, e)
            ws.write(r2, c2 + 6, p)
            ws.write(r2, c2 + 7, fd)
            ws.write(r2, c2 + 8, pp)
            r2+=1

        r2+=2
        ws.write(r2, c2 + 1, "Total")
        ws.write(r2, c2 + 4, sum(reg))
        exd = sum(examd)
        ws.write(r2, c2 + 5, exd)
        psd = sum(passed)
        ws.write(r2, c2 + 6, psd)
        ws.write(r2, c2 + 7, sum(failed))
        ws.write(r2, c2 + 8, round((psd / exd)*100,2))

        wb.close()
        
        output.seek(0)

        response = HttpResponse(
            output,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = "attachment; filename=%s" % wb_name

        return response
            

    return render(request,'admin_template/result_stats.html',context)

@login_required(login_url = 'login')
@allowed_users(allowed_roles=['teacher'])
def add_excel(request):

    c_ss = SemesterSession.objects.filter(active = 'Yes').first()
    print(c_ss)
    print(c_ss.ca_deadline)

    disabled = ""

    from datetime import datetime
    if datetime.now().date() > c_ss.ca_deadline :
        print(True)

        disabled = "disabled"
        messages.success(request,"Deadline to Submit CA Results has passed. You can't submit anymore.")



    t_id = str(request.user.teacher.teacher_id)
    print(t_id)
    data1 = AssignedTeacher2.objects.filter(teacher_id = t_id)

    data = []
    for d in data1:
        if d.course.semester == c_ss.semester :
            data.append(d)

    context={'course':data, 'disabled': disabled} 
    if request.method == 'POST' and request.FILES['myfile']:
        myfile = request.FILES['myfile']
        code = request.POST.get('course_code')
        xx =code.split(",")
        print(xx)
        course_cd = xx[0]
        dept_id =xx[1]

        wb = openpyxl.load_workbook(myfile)
        ws = wb["Sheet1"]
        print(ws)

        excel_data = []

        for r in ws.iter_rows():
            row_data = []
            for cell in r:
                row_data.append(str(cell.value))
            excel_data.append(row_data)
        
        print(excel_data)
        
        for i in range(1, len(excel_data)):
            sub = Result(
                    sem_ses = c_ss,
                    course_code =Subject.objects.filter(course_code = course_cd).first(),
                    theory_marks = excel_data[i][2],
                    dept = dept_id,
                    student_id = excel_data[i][0],
                    level = Student.objects.filter(registration_number = excel_data[i][0]).first().level
                    )
            sub.save()

        messages.success(request,"Successfully Added Results for %s course"% (course_cd))
        return redirect('home') 
    return render(request,'teacher_template/add_json.html',context)

@login_required(login_url = 'login')
@allowed_users(allowed_roles=['teacher'])
def add_exam(request):

    c_ss = SemesterSession.objects.filter(active = 'Yes').first()
    print(c_ss)
    print(c_ss.ca_deadline)

    disabled = ""

    t_id = str(request.user.teacher.teacher_id)
    print(t_id)
    data1 = AssignedTeacher2.objects.filter(teacher_id = t_id)

    data = []
    for d in data1:
        if d.course.semester == c_ss.semester :
            data.append(d)

    context={'course':data, 'disabled': disabled} 
    if request.method == 'POST' and request.FILES['myfile']:
        myfile = request.FILES['myfile']
        code = request.POST.get('course_code')
        xx =code.split(",")
        print(xx)
        course_cd = xx[0]
        dept_id =xx[1]

        wb = openpyxl.load_workbook(myfile)
        ws = wb["Sheet1"]
        print(ws)

        excel_data = []

        for r in ws.iter_rows():
            row_data = []
            for cell in r:
                row_data.append(str(cell.value))
            excel_data.append(row_data)
        
        print(excel_data)
        
        for i in range(1, len(excel_data)):
            stud = ExamCode.objects.filter(code = excel_data[i][0], sem_ses = c_ss, subject = course_cd).first()

            res = Result.objects.filter(sem_ses = c_ss, student = stud.student, course_code = course_cd).first()
            res.term_test = excel_data[i][1]
            res.save()

        messages.success(request,"Successfully Added Results for %s course"% (course_cd))
        return redirect('home') 
    return render(request,'teacher_template/add_exam.html',context)


@login_required(login_url = 'login')
@allowed_users(allowed_roles=['admin'])         
def add_students(request):
    data = Dept.objects.all()
    context={'course':data} 
    if request.method == 'POST' and request.FILES['myfile']:
        myfile = request.FILES['myfile']
        dept_id = request.POST.get('dept')
        dept = Dept.objects.filter(dept_id = dept_id).first()

        print(dept.dept_id)

        wb = openpyxl.load_workbook(myfile)
        ws = wb["Sheet1"]
        print(ws)

        excel_data = []

        for r in ws.iter_rows():
            row_data = []
            for cell in r:
                row_data.append(str(cell.value))
            excel_data.append(row_data)
        
        print(excel_data)
        
        for i in range(1, len(excel_data)):
            user = User(
                username = excel_data[i][0],
                email = excel_data[i][1],
                password = excel_data[i][2],
            )
            user.save()
            
            import pandas as pd
            stud = Student(
                    user = user,
                    registration_number = excel_data[i][5],
                    dept = dept,
                    name = excel_data[i][3],
                    phone = excel_data[i][4],
                    level = excel_data[i][8],
                    dob = pd.to_datetime(excel_data[i][6]).date(),
                    pob = excel_data[i][7],
                )
            stud.save()

        messages.success(request,"Successfully Added Students for %s Department"% (dept_id))
        return redirect('home') 
    return render(request,'student_template/add_students.html',context)  


@login_required(login_url = 'login')
@allowed_users(allowed_roles=['admin'])
def add_admin(request):
    return redirect('register')

@login_required(login_url = 'login')
@allowed_users(allowed_roles=['admin'])
def addDept(request):
    form = DepartmentForm(request.POST or None)
    context = {'form': form, 'page_title':'add department'}
    if request.method == 'POST':
        if form.is_valid:
            form.save()
            messages.success(request,"Successfully Added Dept. ")

    return render(request, 'registration_template/add_dept.html',context)

@login_required(login_url = 'login')
@allowed_users(allowed_roles=['admin'])
def addSchool(request):
    form = SchoolForm(request.POST or None)
    context = {'form': form, 'page_title':'add school'}
    if request.method == 'POST':
        if form.is_valid:
            form.save()
            messages.success(request,"Successfully Added School")

    return render(request, 'registration_template/add_school.html',context)

@login_required(login_url = 'login')
@allowed_users(allowed_roles=['admin'])
def addDegree(request):
    form = DegreeForm(request.POST or None)
    context = {'form': form, 'page_title':'add degree'}
    if request.method == 'POST':
        if form.is_valid:
            form.save()
            messages.success(request,"Successfully Added Degree")

    return render(request, 'registration_template/add_deg.html',context)

@login_required(login_url = 'login')
@allowed_users(allowed_roles=['admin'])
def addSS(request):
    form = SSForm(request.POST or None)
    context = {'form': form, 'page_title':'add semester session'}
    if request.method == 'POST':
        if form.is_valid:
            form.save()
            messages.success(request,"Successfully Added Semester Session")

    return render(request, 'registration_template/add_ss.html',context)

@login_required(login_url = 'login')
@allowed_users(allowed_roles=['admin'])
def setActiveSS(request):
    data = SemesterSession.objects.all()
    print(data)

    active = SemesterSession.objects.filter(active = 'Yes').first()
    print(active)

    context = {'ss' : data, 'current': active}

    if request.method == 'POST':
        sems = SemesterSession.objects.all()
        for s in sems:
            s.active = 'No'
            s.save()
        
        ssid = request.POST.get('ssid')
        print(SemesterSession.objects.filter(ss_id = ssid).first())
        sem = SemesterSession.objects.filter(ss_id = ssid).first()
        print(sem)
        sem.active = 'Yes'
        sem.save()
        messages.success(request,"Semester:  " + sem.ss_id + " is set active.")
    
        print(SemesterSession.objects.filter(active= "Yes").all())

        return redirect('home')
    
    return render(request,'admin_template/set_active_ss.html', context)

@login_required(login_url = 'login')
@allowed_users(allowed_roles=['admin'])
def assign_teacher_dept_search(request):
    data = Result.objects.raw('''
        SELECT dept_id as id FROM main_dept''')

    context={'dept':data} 
    if request.method == 'POST':
        dept_id = request.POST.get('dept_id')
        return redirect(reverse('assign_teacher', kwargs={"dept_id": dept_id}))

    return render(request,'admin_template/assign_teacher_dept_search.html',context)

def assign_teacher(request, dept_id):
    data = Teacher.objects.all()
    course = Subject.objects.filter(dept = dept_id)
    context = {'teacher':data, 'course': course, 'teacher_dept':dept_id}

    if request.method == "POST":

        cour_code = request.POST.get('course')
        t_id   = request.POST.get('teacher')
        t_dept = dept_id

        exists = AssignedTeacher2.objects.filter(course = Subject.objects.filter(course_code = cour_code).first(), teacher_id = t_id).first()

        print(exists)
        if exists:
            messages.success(request,"Teacher already assigned to this course")
        else:

            assingn = AssignedTeacher2(
                course = Subject.objects.filter(course_code = cour_code).first(),
                dept_id = t_dept,
                
                teacher_id= t_id,

            )
            assingn.save()
            messages.success(request,"Teacher Id : %s Is Assigned For %s Course In %s Department" %(t_id,cour_code,t_dept))

    return render(request,'admin_template/assign_teacher_dept.html',context)

@login_required(login_url = 'login')
@allowed_users(allowed_roles=['admin'])
def change_stud_dept(request):
    data1 = Student.objects.all()
    print(data1)
    
    data2 = Dept.objects.all()
    print(data2)


    context = {'studs' : data1, 'depts': data2}

    if request.method == 'POST':
        stud = request.POST.get('stud')
        dept = request.POST.get('dept')

        d = Dept.objects.filter(dept_id = dept).first()
        Student.objects.filter(registration_number = stud).update(dept = d)
        #student.dept = d
        #student.save()
        messages.success(request,"Student:  " + stud + " transfered to " + dept + " Department")
    
    return render(request,'admin_template/change_stud_dept.html', context)

@login_required(login_url = 'login')
@allowed_users(allowed_roles=['admin'])
def promote_stud(request):
    data1 = Student.objects.all()
    print(data1)
    
    data2 = ['HND1','HND2','BTECH']
    print(data2)


    context = {'studs' : data1, 'levs': data2}

    if request.method == 'POST':
        stud = request.POST.get('stud')
        lev = request.POST.get('lev')

        if lev == 'BTECH':
            deg = Degree.objects.filter(deg_id = lev).first()
            Student.objects.filter(registration_number = stud).update(degree_pursued = deg)
            Student.objects.filter(registration_number = stud).update(level = lev)
        else :
            deg = Degree.objects.filter(deg_id = "HND").first()
            Student.objects.filter(registration_number = stud).update(degree_pursued = deg)
            Student.objects.filter(registration_number = stud).update(level = lev)
        
        messages.success(request,"Student:  " + stud + " promoted to " + lev + " Level")
    
    return render(request,'admin_template/promote_stud.html', context)


#--------------------------------------------------------###### ADD END #######---------------------------------------------------------



def student_sub_register(request):
    dept_name = request.user.student.dept
    dpt_name =str(dept_name)
    regi = str(request.user.student.registration_number)
    sem = SemesterSession.objects.filter(active = 'Yes').first()

    print(sem)
    print(sem.semester)
    lev = str(request.user.student.level)

    print(lev)

    data = Subject.objects.filter(dept= dept_name, level = lev, semester = sem.semester)


    for i in data:
        ctt = RegisterTable.objects.filter(subject_id = i.course_code, student_id = regi).first()
        
        if ctt != None:
            i.subject_name = i.subject_name + "-->Already Registered."
            
    context = {'data':data}

    if request.method == "POST":
        course_cc = request.POST.get('course_regi')
        regi = str(request.user.student.registration_number)

        check = RegisterTable.objects.filter(student_id = regi, subject_id = course_cc).first()
        if check == None:
            print("kk")
            
            ss = RegisterTable(
                
                dept_id = dpt_name,
                student_id = regi,
                subject_id = course_cc,


            )
            ss.save()

            tid = AssignedTeacher2.objects.filter(course = course_cc).first().teacher.teacher_id

            print(tid)

            rr = Rating.objects.filter(student_id = regi , subject_id = course_cc, teacher_id = tid).first()
            if rr == None:
                rate = Rating(
                    student_id = regi,
                    subject_id = course_cc,
                    teacher_id = tid,
                )
                rate.save()
            messages.success(request, "Registration Successful")
            return redirect('home')
        else:
            messages.success(request, "You have already Requested for approval on this Subject. See your Registration table for Registration Status")   
    
    return render(request,'teacher_template/student_sub_register.html',context)


def teacher_approve_search(request):
    teach_id = request.user.teacher.teacher_id
    data = AssignedTeacher2.objects.filter(teacher_id = teach_id)

    context= {'data':data}
    if request.method == "POST":
        course_code_dept = request.POST.get('course_code_dept')
        xx = course_code_dept.split(',')
        return redirect(reverse('teacher_approval', kwargs= {"course_code": xx[0], "student_dept": xx[1]}))

    return render(request, 'teacher_template/teacher_approve_search.html',context)


def teacher_approval(request, course_code, student_dept):
    data2 = Result.objects.raw('''
        SELECT 1 as id, name, student_id, phone, status, profile_pic  FROM
        public.main_student JOIN public.main_registertable ON
        main_student.registration_number = main_registertable.student_id
	    where main_registertable.dept_id = %s and main_registertable.subject_id = %s''',[student_dept,course_code]) 
    data = RegisterTable.objects.filter(subject_id = course_code, dept_id = student_dept)
    stu = []
    for i in data:
        stu.append( Student.objects.filter(registration_number = i.student_id).first())

    
        
    context ={'data':data2,'stu':stu, 'cc': course_code, 'dpt': student_dept}
    if request.method == "POST":
        stat = request.POST.get('optionsRadios')
        xx = stat.split(',')
        print(xx)
        if xx[0]=='Pending' or xx[0]=='Rejected':
            Result.objects.filter(student_id = xx[1], course_code = course_code, dept = student_dept).delete()

        RegisterTable.objects.filter(subject_id = course_code, student_id = xx[1]).update(status = xx[0])
        
        tid = str(request.user.teacher.teacher_id)

        if xx[0] == 'Approved':
            rr = Rating.objects.filter(student_id = xx[1], subject_id = course_code, teacher_id = tid).first()
            if rr == None:
                rate = Rating(
                    student_id = xx[1],
                    subject_id = course_code,
                    teacher_id = tid,
                )
                rate.save()

    return render(request, 'teacher_template/teacher_approval.html',context)


def student_rating(request):
    regi =str(request.user.student.registration_number)

    data2 = Result.objects.raw('''
        SELECT 1 as id, name, main_assignedteacher2.teacher_id as tid ,course_code cc, phone, profile_pic,rating  FROM
        public.main_registertable JOIN public.main_assignedteacher2 ON
        main_registertable.dept_id = main_assignedteacher2.student_dept
        AND main_registertable.subject_id = main_assignedteacher2.course_code
        JOIN main_teacher ON main_teacher.teacher_id = main_assignedteacher2.teacher_id
        JOIN main_rating ON main_registertable.student_id = main_rating.student_id 
        AND main_registertable.subject_id = main_rating.subject_id

	    where main_registertable.student_id = %s and main_registertable.status ='Approved'  ''',[regi]) 


    if request.method == "POST":
        stat = request.POST.get('optionsRadios')
        xx = stat.split(',')
        print(xx)
        Rating.objects.filter(subject_id = xx[2], student_id = regi,teacher_id = xx[1]).update(rating = int(xx[0]))

    context = { 'regi':regi,
                'data':data2,

    }

    return render(request, 'student_template/student_rating.html',context)


def get_ratings_teacher(request):
    t_id = str(request.user.teacher.teacher_id)
    attendance = Result.objects.raw('''
    SELECT 1 as id, subject_id as sn , AVG(rating) as avg FROM
    public.main_rating
	where teacher_id=%s
	group by subject_id;''',[t_id])
    data =[]
    labels =[]

    for i in attendance:
        labels.append(i.sn)
        data.append(round(i.avg,2))

    return JsonResponse(data={
        'labels': labels,
        'data':data,
    })

def get_ratings_admin(request):
    t_rate = Result.objects.raw('''
    SELECT 1 as id, main_rating.teacher_id as tid ,main_rating.teacher_id as nem , AVG(rating) as avg FROM
    public.main_rating JOIN main_teacher ON main_rating.teacher_id = main_teacher.teacher_id
	group by main_rating.teacher_id ORDER BY AVG(rating) DESC;''')

    data =[]
    labels =[]

    for i in t_rate:
        nem = Teacher.objects.get(teacher_id = i.tid).name
        dep = str(Teacher.objects.get(teacher_id = i.tid).dept_id)
        i.nem = nem
        labels.append(i.nem + " ["+ i.tid+" ("+ dep + ") "+"]")
        data.append(round(i.avg,2))
    
    return JsonResponse(data={
        'labels': labels,
        'data':data,
    })

def dept_performance(request):
    dep_per = Result.objects.raw('''
    SELECT 1 as id, AVG(total) as avg, dept FROM
    public.main_result 
	group by dept ORDER BY AVG(total) DESC;''')
    data =[]
    labels =[]
    for i in dep_per:
        data.append((i.avg/100)*10)
        labels.append(i.dept)
    return JsonResponse(data={
        'labels': labels,
        'data':data,
    })


@login_required(login_url = 'login')
def subject_ranksheet_teacher(request):
    t_id = request.user.teacher.teacher_id
        
    data = AssignedTeacher2.objects.filter(teacher_id = t_id)

    context={'course':data, 't_id': t_id} 

    c_ss = SemesterSession.objects.filter(active = 'Yes').first()

    if request.method == 'POST':
        course = request.POST.get('course_code')
        xx = course.split(",")
        course_id = xx[0]
        dept_id = xx[1]
        marksObj = Result.objects.filter(course_code = course_id, dept = dept_id, sem_ses = c_ss.ss_id)
        cnt=1
        for i in marksObj:
            i.id = cnt
            cnt = cnt+1
            i.cgname = cal_grade(i.total) 
            print(i.cgname)

        subject = Subject.objects.get(course_code = course_id)
        subject_name = subject.subject_name
        semester = subject.semester
        subject_dept = subject.dept_id
        context={'data':marksObj,'course_id':course_id,'subject_name':subject_name,'session':semester,
                'subject_dept':subject_dept, 'student_dept': dept_id
        } 
        return render(request, 'teacher_template/course_result2.html',context)
    return render(request, 'teacher_template/course_result.html',context)

#______________________________Extract Results____________________________________
@login_required(login_url = 'login')
def extract_results(request):
    
    data = Dept.objects.all()

    context={'dept':data, 'level': ['HND1', 'HND2', 'BTECH']} 

    if request.method == 'POST':
        dept_id = request.POST.get('dept')
        level = request.POST.get('level')

        if level == "HND1":
            lev = 'L1'
        elif level == "HND2":
            lev = 'L2'
        elif level == "BTECH":
            lev = 'L3'


        print(dept_id)
        print(lev)

        matricules = []
        names = []
        dobs = []
        pobs = []
        fields = ["Matricule", "Names", "Date of Birth", "Place of Birth"]
        f_results = []

        students = Student.objects.filter(dept = dept_id, level = lev).all()
        print(students)
        for r in students:
            matricules.append(r.registration_number)

            names.append(r.name)
            dobs.append(r.dob)
            pobs.append(r.pob)
        
        subjects = Subject.objects.filter(dept = dept_id, level = lev).all()
        print(subjects)
        for s in subjects:
            fields.append(s.subject_name)

            res = []
            results = Result.objects.filter(course_code = s.course_code).all()
            for r in results:
                if level == "BTECH":
                    res.append(r.total*5)
                else:
                    res.append(r.total)

            f_results.append(res)
        
        print(matricules)
        print(names)
        print(fields)
        print(f_results)
            

        wb_name = 'Results_' + str(dept_id) + '_' + str(level) + '.xlsx'

        output = io.BytesIO()
        wb = xlsxwriter.Workbook(output)
        ws = wb.add_worksheet('Sheet1')
        
        r1 = 0
        c1 = 0

        for f in fields:
            ws.write(r1, c1, f)
            c1+=1
        r2 = 1
        c2 = 0

        import itertools
        for (m, n, d, p ) in zip(matricules, names, dobs, pobs):
            ws.write(r2, c2, str(m))
            ws.write(r2,c2 + 1, str(n))
            ws.write(r2,c2 + 2, str(d))
            ws.write(r2,c2 + 3, str(p))
            r2+=1

        
        c3 = 4
        for i in range(0, len(f_results)):
            r3 = 1
            for j in range(0, len(matricules)):
                ws.write(r3,c3, f_results[i][j])
                r3+=1
            c3+=1
        


        wb.close()
        
        output.seek(0)

        response = HttpResponse(
            output,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = "attachment; filename=%s" % wb_name

        return response

    return render(request,'admin_template/extract_results.html',context)

#__________________________________________________EXTRACT RESULTS END_____________________________________________________

@login_required(login_url = 'login')
@allowed_users(allowed_roles=['admin'])
def generate_codes(request):

    import random

    sem = SemesterSession.objects.filter(active = 'Yes').first()

    data = Subject.objects.filter(semester = sem.semester)

    context = {'course': data}

    if request.method == 'POST':
        course_code = request.POST.get('course_code').split(",")[0]
        print(course_code)
        my_list = list(range(1,301))
        print(my_list)

        studs = RegisterTable.objects.filter(subject = course_code).all()
        print(studs)

        for stud in studs:
            code = random.choice(my_list)
            my_list.remove(code)

            exam_code = ExamCode(
                sem_ses = sem,
                subject = Subject.objects.filter(course_code = course_code).first(),
                student = stud.student,
                code = code
            )

            exam_code.save()
            
        messages.success(request,"Successfully Generated Codes For Course ")

    return render(request,'admin_template/generate_codes.html',context)

@login_required(login_url = 'login')
@allowed_users(allowed_roles=['admin'])
def download_codes(request):

    sem = SemesterSession.objects.filter(active = 'Yes').first()

    data = Subject.objects.filter(semester = sem.semester)

    context = {'course': data}

    if request.method == 'POST':
        course_code = request.POST.get('course_code').split(",")[0]
        print(course_code)

        course_name = Subject.objects.filter(course_code = course_code).first().subject_name

        matricules = []
        names = []
        codes = []
        fields = ["Matricule", "Names", "Code"]

        register1 = RegisterTable.objects.filter(subject_id = course_code).all()
        for r in register1:
            matricules.append(r.student)

            n = Student.objects.filter(registration_number = r.student).first().name
            names.append(n)

            c = ExamCode.objects.filter(student = r.student, sem_ses = sem, subject = r.subject).first().code
            codes.append(c)

        wb_name = 'Codes_' + str(course_name) + '_' + str(sem.ss_id) +'.xlsx'

        output = io.BytesIO()
        wb = xlsxwriter.Workbook(output)
        ws = wb.add_worksheet('Sheet1')
        
        r1 = 0
        c1 = 0

        for f in fields:
            ws.write(r1, c1, f)
            c1+=1
        r2 = 1
        c2 = 0

        import itertools
        for (m, n, c) in zip(matricules, names, codes):
            ws.write(r2, c2, str(m))
            ws.write(r2,c2 + 1, str(n))
            ws.write(r2,c2 + 2, str(c))
            r2+=1

        wb.close()
        
        output.seek(0)

        response = HttpResponse(
            output,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = "attachment; filename=%s" % wb_name

        return response
            

    return render(request,'admin_template/download_codes.html',context)


def delete_result(request):
    t_id = str(request.user.teacher.teacher_id)
    data = AssignedTeacher2.objects.filter(teacher_id = t_id)
    
    context = {'data':data}
    if request.method == 'POST':
        course_id = request.POST.get('course_code')
        xx = course_id.split(",")
        print(xx)
        
        return redirect(reverse('delete_result2', kwargs= {"dept": xx[1], "course_id": xx[0]}))

    return render(request,'teacher_template/delete_result.html',context)


def delete_result2(request, dept, course_id):
    data = Result.objects.filter(dept = dept, course_code = course_id)

    if request.method == 'POST':
        course_id = request.POST.get('course_code')
        xx = course_id.split(",")

        Result.objects.filter(course_code = xx[1], student_id = xx[0]).delete()

        messages.success(request,"Successfully Deleted %s student's %s Course Result"%(xx[0],xx[1]))
        return redirect('home')
    
    return render(request,'teacher_template/delete_result.html')

def delete_student(request):
    data = Student.objects.all()

    if request.method == 'POST':
        regi = request.POST.get('course_code')
    
        uid = Student.objects.get(registration_number = regi).user_id
        Student.objects.filter(registration_number = regi).delete()
        User.objects.filter(id = uid).delete()

        messages.success(request,"Successfully Deleted %s student"%(regi))
        return redirect('home')

    context ={
        'data':data


    }
    return render(request,'admin_template/remove_student.html',context)

def remove_teacher(request):
    data = Teacher.objects.all()

    if request.method == 'POST':
        regi = request.POST.get('course_code')
    
        uid = Teacher.objects.get(teacher_id = regi).user_id
        Teacher.objects.filter(teacher_id = regi).delete()
        User.objects.filter(id = uid).delete()

        messages.success(request,"Successfully Deleted %s Teacher"%(regi))
        return redirect('home')



    context ={
        'data':data


    }
    return render(request,'admin_template/remove_teacher.html',context)

class GeneratePdf2(View):
    course_id = None
    dept_id = None
    def get(self, request, *args, **kwargs):
        self.course_id = self.kwargs.get('course_id', None)
        self.dept_id = self.kwargs.get('dept_id',None)

        cid = str(self.course_id)
        did = str(self.dept_id)
        print(cid)
        print(did)
        marksObj = Result.objects.filter(course_code = cid, dept = did, sem_ses = SemesterSession.objects.filter(active= 'Yes').first())
        cnt=1
        for i in marksObj:
            i.id = cnt
            cnt = cnt+1
            i.cgname = cal_grade(i.total) 
            # print(i.cgname)

    
        dept_name = Dept.objects.get(dept_id = did).name
        subject_name = Subject.objects.get(course_code = cid).subject_name 
        credit = Subject.objects.get(course_code = cid).credit 
        session  = Subject.objects.get(course_code = cid).semester

        module_dir = os.path.dirname(__file__)  # get current directory
        file_path1 = os.path.join(module_dir, 'templates/teacher_template/generate_result_pdf_temp.html')
        file_path2 = os.path.join(module_dir, 'templates/teacher_template/generate_result_pdf.html')

        pwd = os.path.dirname(__file__)
        open(file_path1, "w").write(render_to_string(file_path2, {'data': marksObj, 'course_number': cid,'dept_name': dept_name,'subject_name': subject_name,'credit':credit ,'session':session}))

        # Converting the HTML template into a PDF file
        pdf = html_to_pdf(file_path1)

        # rendering the template
        return HttpResponse(pdf, content_type='application/pdf')

