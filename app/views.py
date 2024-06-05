import json
import xlsxwriter
import io
import openpyxl
from django.contrib.auth.models import Group
from django.contrib import messages
from django.core.files.storage import FileSystemStorage
from django.http import HttpResponse, JsonResponse
from django.shortcuts import (HttpResponse,
                              get_object_or_404, redirect, render)
from django.urls import reverse
from django.contrib.auth.decorators import login_required

from .models import *
from .forms import *

from django.contrib.auth import authenticate, login, logout
from .decorators import allowed_users, unauthenticated_user,allowed_users_home
from django.http import HttpResponse
from django.views.generic import View

import io
import os
from django.db.models import Avg

from django.db.utils import *

from django.contrib.auth.models import User


def cal_cg(marks):
    if marks>=16:
        return 4.0
    elif marks>=14:
        return 3.5
    elif marks>=12:
        return 3.00
    elif marks>=11:
        return 2.50
    elif marks>=10:
        return 2.00
    elif marks>=9:
        return 1.50
    elif marks>=8:
        return 1.00
    else:
        return 0.00
    
def cal_grade(mark):
    if mark>=15:
        return "Distinction"
    elif mark>=14:
        return "Upper Credit"
    elif mark>=12:
        return "Lower Credit"
    elif mark>=10:
        return "Pass"
    else:
        return "Fail"
    
def cal_class(mark):
    if mark>=3.6:
        return "First Class"
    elif mark>=3.0:
        return "Second Class - UD"
    elif mark>=2.5:
        return "Second Class - LD"
    elif mark>=2.25:
        return "Third Class"
    elif mark>=2.0:
        return "Pass"
    else:
        return "Fail"

def cal_cgname(cg):
    if cg == 4.00:
        return "A"
    elif cg == 3.50:
        return "B+"
    elif cg == 3.00:
        return "B"
    elif cg == 2.50:
        return "C+"
    elif cg == 2.00:
        return "C"
    elif cg == 1.50:
        return "D+"
    elif cg == 1.00:
        return "D"
    elif cg == 0.00:
        return "F"


##-----------------------------LOG IN AND REGISTER--------------------------------------------------------###

@unauthenticated_user
def loginPage(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(request, username=username, password=password)
        

        if user is not None:
            login(request,user)
            if user.is_active :
                return redirect('home')
            else:
                messages.info(request, "User blocked")
            
        else:
            messages.info(request, "Username Or Password is not Correct or User is blocked.")
        

    return render(request, 'login_template/login1.html')

@login_required(login_url = 'login')
@allowed_users(allowed_roles=['admin'])
def registerPageTeacher(request):
    user_form = CreateUserForm(request.POST or None) 
    teacher_form = TeacherForm(request.POST or None, request.FILES or None)
    context = {'teacher_form': teacher_form,'user_form':user_form, 'page_title':'Add Teacher'}
    if request.method == 'POST':
        if user_form.is_valid and teacher_form.is_valid():
            user = user_form.save()
            teacher = teacher_form.save()
            teacher.user =user
            teacher.save()
            
            group = Group.objects.get(name = 'teacher')
            user.groups.add(group)
            messages.success(request, "Successfully Added Teacher")
        else:
            messages.success(request, "Form not Valid. Try another username and/or Lecturer ID and fill all fields correctly.")
            
            
    return render(request, 'registration_template/add_teacher.html',context)

@login_required(login_url = 'login')
@allowed_users(allowed_roles=['admin'])
def registerPage(request):
    user_form = CreateUserForm(request.POST or None) 
    admin_form = AdminForm(request.POST or None, request.FILES or None)
    context = {'admin_form': admin_form,'user_form':user_form, 'page_title':'Add Admin'}
    if request.method == 'POST':
        if user_form.is_valid() and admin_form.is_valid():
            user = user_form.save()
            admin = admin_form.save()
            admin.user =user
            admin.save()
            
            group = Group.objects.get(name = 'admin')
            user.groups.add(group)
            messages.success(request, "Successfully Added Admin")
        else:
            messages.success(request, "Form not valid. Try another username and enter all fields correctly please.")
    return render(request, 'registration_template/add_admin.html',context)

def logoutPage(request):
    logout(request)
    return redirect('login') 

##-----------------------------------------------LOG IN END -------------------------------------------------------###


##----------------------------------------------- HOME PAGE ----------------------------------------------------------##
@login_required(login_url = 'login')
@allowed_users_home(allowed_roles=['admin', 'teacher'])
def home(request):
    name = str(request.user.adminuser.name)
    stu_cnt = Student.objects.all().count()
    teacher_cnt =Teacher.objects.all().count()
    school_cnt = School.objects.all().count()
    admin_cnt = AdminUser.objects.all().count()
    sub_cnt = Subject.objects.all().count()
    overall_rate =Rating.objects.aggregate(Avg('rating'))
    hi = Result.objects.raw ('''
    SELECT 1 as id,dept, AVG(total) as avg from app_result group by dept ORDER BY avg DESC  LIMIT 1;
    
    ''')
    low = Result.objects.raw ('''
    SELECT 1 as id,dept, AVG(total) as avg from app_result group by dept ORDER BY avg LIMIT 1;
    
    ''')
    hi_dept =""
    low_dept =""
    hi_mark =0
    low_mark =0
    for i in hi:
        hi_dept = i.dept
        hi_mark = round((i.avg)/4.0, 2)
    for i in low:
        low_dept = i.dept
        low_mark = round((i.avg)/4.0, 2)

    
    if overall_rate['rating__avg'] is None:
        overall_rate = {'rating__avg': 0 }
    

    context = {'name':name,
                'stu':stu_cnt,
                'teacher':teacher_cnt,
                'school': school_cnt,
                'admin_cnt': admin_cnt,
                'sub_cnt':sub_cnt,
                'overall_rate': round(overall_rate['rating__avg'],2),
                'hi_dept': hi_dept,
                'hi_mark':hi_mark,
                'low_dept': low_dept,
                'low_mark':low_mark,                

    }
    if request.method == 'POST':
        # Register all students to various courses for the semester
        sem = SemesterSession.objects.filter(active= "Yes").first()
        from django.core.paginator import Paginator

        queryset = Student.objects.select_related('degree_pursued','dept').all().order_by('registration_number')
        paginator = Paginator(queryset, per_page=50) 
        try:  
            # Iterate through pages
            for page_num in range(1, paginator.num_pages + 1):
                page = paginator.page(page_num)
                for stud in page.object_list:
                    deg = stud.degree_pursued.deg_id
                    lev = stud.level
                    dep = stud.dept

                    if deg == "HND":
                        sub = "HND"
                    elif deg == "BTECH":
                        sub = "DEGREE"
                    else:
                        sub = "MASTER"

                    courses = Subject.objects.filter(level = lev, subtype = sub, dept= dep, semester = sem.semester)

                    for c in courses:
                        check = RegisterTable.objects.filter(student = stud, subject = c, sem_ses = sem).first()
                        if check == None:
                            print('registering student: ' + str(stud) + ' to course: ' + str(c.subject_name))
                            ss = RegisterTable(
                                sem_ses = sem,
                                dept = dep,
                                student = stud,
                                subject = c
                            )
                            ss.save()
                            
                            if AssignedTeacher2.objects.filter(course_id = c.course_code).first() != None:

                                tid = AssignedTeacher2.objects.filter(course_id = c.course_code).first().teacher.teacher_id

                                rr = Rating.objects.filter(student = stud , subject_id = c.course_code, teacher_id = tid).first()
                                if rr == None:
                                    rate = Rating(
                                        student = stud,
                                        subject_id = c.course_code,
                                        teacher_id = tid,
                                    )
                                    rate.save()
                                else:
                                    print('Continuing loop because rating exist...')
                                    continue
                            else: 
                                print('Continuing loop because no teacher is assigned to course...')
                                continue
                        else:
                            print('Continuing loop because student already registered...')
                            continue
            
            messages.success(request,"All students registered to respective courses.")

        except Exception as e: 
            messages.error(request,'Error: '+ str(e))   

        return redirect('home')
    
    return render(request,'admin_template/index.html',context)

@login_required(login_url = 'login')
@allowed_users(allowed_roles=['student'])
def studentHome(request):
    name = request.user.student.name
    regi = request.user.student.registration_number
    dept = request.user.student.dept
    res = RegisterTable.objects.filter(student_id = regi)
    creds = request.user.student.degree_pursued.total_credits
    deg = request.user.student.degree_pursued.deg_id
    c_ss = SemesterSession.objects.filter(active = 'Yes').first()
    pe = 'auto'

    if c_ss.results_published == "No":
        pe = 'none'

    if res == None:
        data = []
        credits = 0
        credits_passed = 0
        percent_passed_credit = 0
        percent_registered = 0
        remain_credit = creds
        remain_credit_percent= 100
        current_cgpa =0
        degree_status = "INCOMPLETE"
        remain = "Need to pass "+ str(remain_credit) +" more credits to get a degree" 
    
    else:
        credits = 0
        for r in res:
            sub = Subject.objects.filter(course_code = r.subject.course_code).first()
            credits += sub.credit

        credits_passed = 0
        res = Result.objects.filter(student_id = regi).all()

        for r in res:
            if r.total >= 10:
                credits_passed+= r.course_code.credit

        degree_status = ""


        upper =0
        lower = 0
        for r in res:
            if deg == "HND":
                upper =upper+ r.total
            else:
                upper += cal_cg(r.total)
            lower += 1

        if lower == 0:
            lower = 1
        
        if credits_passed is None:
            credits_passed = 0

        current_cgpa = round(upper/lower,2)
        percent_passed_credit = ((credits_passed)*100)/creds


        percent_registered = ((credits)*100)/creds

        remain_credit = creds - credits_passed

        remain_credit_percent = ((remain_credit)*100)/creds
        remain =""
        if(credits_passed<creds):
            degree_status = "INCOMPLETE"
            remain = "Need to pass "+ str(remain_credit) +" more credits to get a degree" 
        else:
            degree_status = "COMPLETE"
            remain = "Congratulations!!! You are a YIBS "+ str(dept) +" Graduate"  

    context ={'name':name, 
    'credits': credits,
    'percent_registered': percent_registered,
    'credits_passed' : credits_passed,
    'percent_passed_credit': percent_passed_credit,
    'current_cgpa': current_cgpa,
    'degree_status': degree_status,
    'remain': remain,
    'creds': creds,
    'pe': pe,
    'aca': c_ss.session 
    }
    return render(request,'student_template/index.html',context)

@login_required(login_url = 'login')
def teacher_home(request):
    name = str(request.user.teacher.name)
    dept = str(request.user.teacher.dept_id)
    t_id = str(request.user.teacher.teacher_id)
    num_of_regi_sub = AssignedTeacher2.objects.filter(teacher_id = t_id).count()
    context ={ 'name':name,
                'dept':dept,
                't_id': t_id,
                'n_ass': num_of_regi_sub,

    }
    return render(request,'teacher_template/index.html',context)

@login_required(login_url = 'login')
def courseWiseParticipation(request, *args, **kwargs):
    t_id = str(request.user.teacher.teacher_id)
    attendance = Result.objects.raw('''
    SELECT 1 as id, app_assignedteacher2.course_code as sn , SUM(attendence) as attend FROM
    public.app_assignedteacher2 JOIN public.app_result ON
    app_assignedteacher2.course_code = app_result.course_code AND student_dept = dept
	where app_assignedteacher2.teacher_id=%s
	group by app_assignedteacher2.course_code;''',[t_id])
    data =[]
    labels =[]
    for i in attendance:
        labels.append(i.sn)
        data.append(i.attend)

    return JsonResponse(data={
        'labels': labels,
        'data':data,
    })


@login_required(login_url = 'login')
def course_wise_performance(request, *args, **kwargs):
    t_id = str(request.user.teacher.teacher_id)
    attendance = Result.objects.raw('''
    SELECT 1 as id, app_assignedteacher2.course_code as sn , SUM(total) as total, COUNT(*) as cnt FROM
    public.app_assignedteacher2 JOIN public.app_result ON
    app_assignedteacher2.course_code = app_result.course_code AND student_dept = dept
	where app_assignedteacher2.teacher_id=%s
	group by app_assignedteacher2.course_code;''',[t_id])
    data =[]
    labels =[]

    for i in attendance:
        labels.append(i.sn)
        data.append((i.total/(i.cnt*100.0))*10)

    return JsonResponse(data={
        'labels': labels,
        'data':data,
    })

@login_required(login_url = 'login')
def session_wise_courses(request, *args, **kwargs):
    t_id = str(request.user.teacher.teacher_id)
    attendance = Result.objects.raw('''
    SELECT 1 as id, semester as sn , COUNT(*) as cnt FROM
    public.app_assignedteacher2
    JOIN app_subject ON app_assignedteacher2.course_code = app_subject.course_code
	where app_assignedteacher2.teacher_id=%s
	group by semester;''',[t_id])
    data =[]
    labels =[]

    for i in attendance:
        labels.append(i.sn)
        data.append(i.cnt)

    return JsonResponse(data={
        'labels': labels,
        'data':data,
    })

#--------------------------------------------*****GET*****-------------------------------------------------------------#

@login_required(login_url = 'login')
def get_att(request, *args, **kwargs):
    regi = request.user.student.registration_number
    attendance = Result.objects.raw('''
    SELECT 1 as id, subject_name as sn , attendence as attend FROM
    public.app_student JOIN public.app_result ON
    app_student.registration_number = app_result.student_id
    JOIN public.app_subject ON app_result.course_code = app_subject.course_code
    where app_student.registration_number=%s;''',[regi])
    data =[]
    labels =[]
    for i in attendance:
        labels.append(i.sn)
        data.append(i.attend)

    return JsonResponse(data={
        'labels': labels,
        'data':data,
    })
@login_required(login_url = 'login')
def get_subtype(request, *args, **kwargs):
    regi = request.user.student.registration_number
    subtype = Result.objects.raw('''
    SELECT 1 as id, subtype, SUM(total) as sum_marks, count(subtype) as cnt FROM
    public.app_student JOIN public.app_result ON
    app_student.registration_number = app_result.student_id
    JOIN public.app_subject ON app_result.course_code = app_subject.course_code
	where registration_number = %s
    group by subtype;''',[regi])

    data =[]
    labels =[]
    for i in subtype:
        labels.append(i.subtype)
        cntt = min(i.cnt*.75,3)
        data.append((i.sum_marks)*8/(i.cnt*100) + cntt)

    return JsonResponse(data={
        'labels': labels,
        'data':data,
    })
def getting_json(subtype, regi):

    marksObj = Result.objects.raw('''
    SELECT 1 as id, subject_name, app_subject.course_code as cc, total FROM
    public.app_student JOIN public.app_result ON
    app_student.registration_number = app_result.student_id
    JOIN public.app_subject ON app_result.course_code = app_subject.course_code
	where registration_number = %s and subtype = %s;''',[regi,subtype])
    

    subject_name =[]
    course_code =[]
    marks = []
    attr = []
    attr.append("subject_name")
    attr.append("course_code")
    attr.append("marks")
    json_res =[]
    for i in marksObj:
        obj = {}
        obj[attr[0]] = i.subject_name
        obj[attr[1]] = i.cc
        
        cg = cal_cg(i.total)
        obj[attr[2]]= cg
        json_res.append(obj) 
    return json_res


def getting_json_result(regi):

    marksObj = Result.objects.filter(student = regi)

    c_ss = SemesterSession.objects.filter(active = 'Yes').first()
    
    pub_marks = []

    if c_ss.results_published == "No":
        print('exams not published')
        for m in marksObj:
            print(m.sem_ses)
            if str(m.sem_ses) != str(c_ss.ss_id):
                pub_marks.append(m)
    else:
        pub_marks = marksObj
    attr = []
    
    attr.append("subject_name")
    attr.append("course_code")
    attr.append("credit")
    attr.append("theory")
    attr.append("tt")
    attr.append("total")
    attr.append("cgpa")
    json_res =[]
    for i in pub_marks:
        obj = {}
        obj[attr[0]] = i.course_code.subject_name
        obj[attr[1]] = i.course_code.course_code
        obj[attr[2]]= i.course_code.credit
        obj[attr[3]] = i.theory_marks
        obj[attr[4]] = i.term_test
        if regi.level == "HND1" or regi.level == "HND2":
            obj[attr[5]] = i.total
            cgpa = cal_grade(i.total) 
        else:
            cgpa = cal_cgname(cal_cg(i.total))
            obj[attr[5]] = i.total * 5
            
        obj[attr[6]]= cgpa
        json_res.append(obj) 
        
    return json_res

def getting_students_json():
    Studs = Student.objects.all().order_by('dept', 'name')

    attr = [] 
    attr.append("Matricule")
    attr.append("Name")
    attr.append('Department')
    attr.append("Degree Pursued")
    attr.append("Level")

    json_res = []
    for stud in Studs:
        obj = {}
        obj[attr[0]] = stud.registration_number
        obj[attr[1]] = stud.name
        obj[attr[2]] = stud.dept.dept_id
        obj[attr[3]] = stud.degree_pursued.deg_id
        obj[attr[4]] = stud.level
        json_res.append(obj)
    
    return json_res

def getting_courses_json():
    Subs = Subject.objects.all().order_by('dep', 'level', 'subject_name')

    attr = [] 
    attr.append("Course_Code")
    attr.append("Name")
    attr.append('Department')
    attr.append("Level")
    attr.append("Credit")
    attr.append("Lecturer")

    json_res = []
    for sub in Subs:
        obj = {}
        obj[attr[0]] = sub.course_code
        obj[attr[1]] = sub.subject_name
        obj[attr[2]] = sub.dep
        obj[attr[3]] = sub.level
        obj[attr[4]] = sub.credit
        ass = AssignedTeacher2.objects.filter(course = sub).first()
        if ass is not None:
            obj[attr[5]] = ass.teacher.name
        else:
            obj[attr[5]] = " "
        json_res.append(obj)
    
    return json_res

def getting_lecturers_json():
    Lects = Teacher.objects.all().order_by('name')

    attr = [] 
    attr.append("Name")
    attr.append('N_Courses')
    attr.append("Courses")

    json_res = []
    for lect in Lects:
        obj = {}
        obj[attr[0]] = lect.name
        obj[attr[1]] = AssignedTeacher2.objects.filter(teacher = lect).all().count()
        obj[attr[2]] = ""
        for ass in AssignedTeacher2.objects.filter(teacher = lect).all():
            obj[attr[2]]+= ass.course.subject_name + '<p></p>'
        json_res.append(obj)
    
    return json_res

def getting_schools_json():
    Schools = School.objects.all().order_by('name')

    attr = [] 
    attr.append("Name")
    attr.append('N_Depts')
    attr.append("Depts")

    json_res = []
    for sch in Schools:
        obj = {}
        obj[attr[0]] = sch.name
        obj[attr[1]] = Dept.objects.filter(school = sch).all().count()
        obj[attr[2]] = ""
        for dep in Dept.objects.filter(school = sch).all():
            obj[attr[2]]+= dep.name + '<p></p> '
        json_res.append(obj)
    
    return json_res

def teacher_subject_list(request):


    t_id  = request.user.teacher.teacher_id

    data = AssignedTeacher2.objects.filter(teacher = t_id)

    attr = []
    attr.append("name")
    attr.append("cc")
    attr.append("subtype")
    json_res =[]
    for i in data:
        obj = {}
        obj[attr[0]] = i.course.subject_name
        obj[attr[1]] = i.course.course_code
        obj[attr[2]] = i.course.subtype
        
        
        json_res.append(obj) 
        
    return JsonResponse(json_res, safe = False) 

@login_required(login_url = 'login')
def get_all_the_marks(request, *args, **kwargs):
    regi = request.user.student
    json_res = getting_json_result(regi)
    return JsonResponse(json_res, safe = False)

@login_required(login_url = 'login')
def get_all_the_students(request, *args, **kwargs):
    json_res = getting_students_json()
    return JsonResponse(json_res, safe = False)

@login_required(login_url = 'login')
def get_all_the_courses(request, *args, **kwargs):
    json_res = getting_courses_json()
    return JsonResponse(json_res, safe = False)

@login_required(login_url = 'login')
def get_all_the_lecturers(request, *args, **kwargs):
    json_res = getting_lecturers_json()
    return JsonResponse(json_res, safe = False)

@login_required(login_url = 'login')
def get_all_the_schools(request, *args, **kwargs):
    json_res = getting_schools_json()
    return JsonResponse(json_res, safe = False)

def see_registration_status(request, *args, **kwargs):
    regi = str(request.user.student.registration_number)
    dep = str(request.user.student.dept)
    register = Result.objects.raw('''
    SELECT 1 as id, status, subject_id, dept_id as sub, dept_id as teacher FROM app_registertable
	where student_id = %s ''',[regi])

    for i in register:
        c_id = i.subject_id
        subject_name = Subject.objects.get(course_code = c_id).subject_name  
        
    attr=[]
    attr.append("course_code")
    attr.append("status")
    json_res =[]
    for i in register:
        obj = {}
        obj[attr[0]] = i.subject_id
        obj[attr[1]]= i.status
        json_res.append(obj)
    return JsonResponse(json_res, safe= False)
    

#------------------------------------------------------***VIEW***-----------------------------------------------------------
@login_required(login_url = 'login')
def full_attendance(request):
    return render(request,'student_template/full_attendance.html')

@login_required(login_url = 'login')
def full_marksheet(request):
    c_ss = SemesterSession.objects.filter(active = 'Yes').first()
    pe = "auto"
    if c_ss.results_published == "No":
        pe = "none"
    return render(request,'student_template/full_marksheet.html', {'pe': pe, 'aca': c_ss.session})

@login_required(login_url = 'login')
def all_students(request):
    return render(request,'admin_template/view_studs.html')

@login_required(login_url = 'login')
def all_lecturers(request):
    return render(request,'admin_template/view_lecturers.html')

@login_required(login_url = 'login')
def all_courses(request):
    return render(request,'admin_template/view_courses.html')

@login_required(login_url = 'login')
def all_schools(request):
    return render(request,'admin_template/view_schools.html')

@login_required(login_url = 'login')
def full_skillset(request):
    return render(request,'student_template/full_skillset.html')

@login_required(login_url = 'login')
def subject_ranksheet(request):
    regi = request.user.student.registration_number
        
    data = Result.objects.raw('''
    SELECT 1 as id, course_code FROM app_result
	where student_id = %s''',[regi])

    context={'course':data, 'regi': regi} 

    if request.method == 'POST':
        course_id = request.POST.get('course_code')
        marksObj = Result.objects.raw('''
        SELECT 1 as id,app_student.name, app_student.registration_number as regi, total as marks FROM
        public.app_student JOIN public.app_result ON
        app_student.registration_number = app_result.student_id
        JOIN public.app_subject ON app_result.course_code = app_subject.course_code
	    where app_subject.course_code = %s order by marks DESC;''',[course_id])
        cnt=1
        your_rank = 0
        regi1 = request.user.student.registration_number
        for i in marksObj:
            if i.regi == regi1:
                your_rank = cnt  
            i.id =cnt
            cnt = cnt+1
        subject_name = Subject.objects.get(course_code = course_id).subject_name
        context={'data':marksObj,'course_id':course_id,'subject_name':subject_name, 'rank':your_rank} 
        return render(request, 'student_template/rank_result2.html',context)


    return render(request, 'student_template/rank_result.html',context)

from django.http import HttpResponse
from django.views.generic import View

from app.utils import html_to_pdf 
from django.template.loader import render_to_string
import os
class GeneratePdf(View):
     def get(self, request, *args, **kwargs):
        regi = request.user.student.registration_number
        name = request.user.student.name
        phone = request.user.student.phone
        email = request.user.email
        dept = request.user.student.dept.name
        deg = request.user.student.degree_pursued.deg_id
        data = Result.objects.filter(student = regi).all()

        upper =0
        count = 0
        lower = 0
        cgpa = 0
        status =""
        cnt = 1
        for k in data:
            if k.total>=10:
                lower =lower + k.course_code.credit
            if deg == "HND":
                upper+=k.total
            else:
                k.gp = cal_cg(k.total)
                upper+= k.gp

            count+=1
            k.id = cnt
            cnt = cnt+1
        
        if lower == 0:
            cgpa =0
            
        else:
            cgpa = round(upper/count, 2)

        status = cal_grade(cgpa)

        if lower < request.user.student.degree_pursued.total_credits:
            status = "Incomplete"
        else:
            status = "Complete"

        for i in data:
            if deg == "HND":
                i.student_id = cal_grade(i.total)
            else:
                i.student_id = cal_cgname(i.gp)
                i.total = i.total * 5    

        module_dir = os.path.dirname(__file__)  # get current directory
        file_path1 = os.path.join(module_dir, 'templates/student_template/generate_result_pdf_temp.html')
        file_path2 = os.path.join(module_dir, 'templates/student_template/generate_result_pdf.html')

        pwd = os.path.dirname(__file__)
        open(file_path1, "w").write(render_to_string(file_path2, {'data': data,'regi':regi,'name':name, 'phone': phone, 'email':email,'cgpa':cgpa,'status':status,'dept':dept }))

        # Converting the HTML template into a PDF file
        pdf = html_to_pdf(file_path1)
        
        # rendering the template
        return HttpResponse(pdf, content_type='application/pdf')

def my_trans(request):
    data = AcademicYear.objects.all()

    context = {'ays': data}

    if request.method == 'POST':
        aca = request.POST.get('ay')
        return redirect(reverse('generate_trans', kwargs={"ay": aca}))

    return render(request, 'student_template/trans.html',context)

def admin_trans(request):
    data = AcademicYear.objects.all()

    depts = Dept.objects.all()

    context = {'ays': data, 'ds': depts}

    if request.method == 'POST':
        aca = request.POST.get('ay')
        dept = request.POST.get('dept')
        return redirect(reverse('generate_trans', kwargs={"ay": aca, "dept": dept}))

    return render(request, 'admin_template/trans.html',context)


class GenerateTrans(View):
    ay = None
    d = None
    studs = []
    def get(self, request, *args, **kwargs):
        self.ay = self.kwargs.get('ay', None)
        self.d = self.kwargs.get('dept', None)

        if self.d != None:
            self.studs = Student.objects.filter(dept = self.d).all()
            zip_name = "Transcripts" + "_" + self.d + "_" + self.ay + ".zip"
        else:
            self.studs.append(request.user.student)
            zip_name = "Transcripts" + "_" + str(self.studs[0].name) + "_" + self.ay + ".zip"
        
        students = self.studs

        head = Image.objects.filter(title = 'trans_head').first()

        pdfs = []
        names = []
        resit_display = False

        for stud in students:

            regi = stud.registration_number
            name = stud.name
            phone = stud.phone
            email = stud.user.email
            dept = stud.dept
            program = stud.degree_pursued
            dob = stud.dob
            pob = stud.pob
            data1 = Result.objects.filter(student = regi).all().order_by('course_code')

            ay = str(self.ay)

            f_name = "Transcript " + "_" + name + "_" + program.deg_id + "_" + ay + ".pdf"

            data = []
            for d in data1:
                if str(d.sem_ses.session) == ay:
                    data.append(d)


            if len(data) == 0:
                messages.success(request,"Student %s has no results for the choosen Academic Year"%(stud.name))
            else:
                resits = []
                sem1 = []
                gpa_sum_s1 = 0
                cred_total_s1 = 0
                cred_earned = 0
                gpa_sem1 = 0
                gp_s1 = 0
                grade_sem1 = ""
                count = 0

                wpt_s1 = 0
                gpt_s1 = 0

                for d in data:
                    if str(d.sem_ses.semester) == 'Semester 1':
                        sem1.append(d)
                
                if len(sem1) > 0:
                    level = sem1[0].level
                    cnt = 1
                    for s in sem1:
                        count+=1
                        gpa_sum_s1+= s.total
                        cred_total_s1+= s.course_code.credit
                        if s.absent == "Yes":
                            s.student_id = "X"
                            s.gp = "X"
                            s.gd = "X"
                            s.wp = "X"
                            s.total = "X"
                        else:
                            s.student_id = cal_grade(s.total)
                            s.id = cnt
                            s.gp = cal_cg(s.total)
                            s.gd = cal_cgname(s.gp)
                            s.wp = s.course_code.credit * s.gp
                            if program.deg_id not in ["HND1", "HND2"]:
                                s.total = s.total * 5 
                        if s.resited == "Yes":
                            resits.append(s)

                        cnt = cnt+1
                        
                        if program.deg_id not in ["HND1", "HND2"] and s.total != "X" and s.total >= 50:
                            cred_earned+= s.course_code.credit

                    
                    for s in sem1:
                        if s.wp != "X"  and s.gp != "X":
                            wpt_s1+= s.wp
                            gpt_s1+= s.gp
                    
                    gpa_sem1 = round(gpa_sum_s1/count, 2)
                    gp_s1 = round(gpt_s1/count, 2)
                    grade_sem1 = cal_grade(gpa_sem1)

                sem2 = []
                gpa_sum_s2 = 0
                cred_total_s2 = 0
                gpa_sem2 = 0
                gp_s2 = 0
                grade_sem2 = ""
                count = 0

                wpt_s2 = 0
                gpt_s2 = 0
                
                for d in data:
                    if str(d.sem_ses.semester) == 'Semester 2':
                        sem2.append(d)
                
                if len(sem2) > 0:
                    cnt = 1
                    for s in sem2:
                        count+=1
                        gpa_sum_s2+= s.total
                        cred_total_s2+= s.course_code.credit
                        if s.absent == "Yes":
                            s.student_id = "X"
                            s.gp = "X"
                            s.gd = "X"
                            s.wp = "X"
                            s.total = "X"
                        else:
                            s.student_id = cal_grade(s.total)
                            s.id = cnt
                            s.gp = cal_cg(s.total)
                            s.gd = cal_cgname(s.gp)
                            s.wp = s.course_code.credit * s.gp
                            s.total = s.total * 5 

                        if s.resited == "Yes":
                                resits.append(s)

                        cnt = cnt+1

                        if program.deg_id not in ["HND1", "HND2"] and s.total != "X" and s.total >=50:
                            cred_earned+= s.course_code.credit

                    for s in sem2:
                        if s.wp != "X"  and s.gp != "X":
                            wpt_s2+= s.wp
                            gpt_s2+= s.gp
                    
                    gpa_sem2 = round(gpa_sum_s2/count, 2)
                    gp_s2 = round(gpt_s2/count, 2)
                    grade_sem2 = cal_grade(gpa_sem2)
                
                gpa_sum_resit = 0
                cred_total_resit = 0
                gpa_resit = 0
                gp_resit = 0
                gp_resit = 0
                grade_resit = ""
                count = 0
                wpt_resit = 0
                gpt_resit = 0
                if len(resits) > 0:
                    resit_display = True
                    cnt = 1
                    for s in resits:
                        count+=1
                        gpa_sum_resit+= s.total_resit
                        cred_total_resit+= s.course_code.credit
                        if s.absent == "Yes":
                            s.student_id2 = "X"
                            s.gp2 = "X"
                            s.gd2 = "X"
                            s.wp2 = "X"
                            s.total_resit = "X"
                        else:
                            s.student_id2 = cal_grade(s.total_resit)
                            s.id2 = cnt
                            s.gp2 = cal_cg(s.total_resit)
                            s.gd2 = cal_cgname(s.gp2)
                            s.wp2 = s.course_code.credit * s.gp2
                            s.total_resit = s.total_resit * 5 
                        cnt = cnt+1
                        
                        if program.deg_id not in ["HND1", "HND2"] and s.total_resit != "X" and s.total_resit >= 50:
                            cred_earned+= s.course_code.credit
                    
                    
                    for s in resits:
                        if s.wp != "X"  and s.gp != "X":
                            wpt_resit+= s.wp2
                            gpt_resit+= s.gp2
                    
                    gpa_resit = round(gpa_sum_resit/count, 2)
                    gp_resit = round(gpt_resit/count, 2)
                    grade_resit = cal_grade(gpa_resit)

                if gpa_resit != 0 and gp_resit != 0:
                    year_gpa = round((gpa_sem1 + gpa_sem2 + gpa_resit)/3, 2)
                    year_grade = cal_grade(year_gpa)
                    yr_gp = round((gp_s1 + gp_s2 + gp_resit)/3, 2)
                elif gpa_sem2 == 0 and gp_s2 == 0:
                    year_gpa = gpa_sem1
                    year_grade = cal_grade(year_gpa)
                    yr_gp = gp_s1
                else:
                    year_gpa = round((gpa_sem1 + gpa_sem2)/2, 2)
                    year_grade = cal_grade(year_gpa)
                    yr_gp = round((gp_s1 + gp_s2)/2, 2)

                cred_attempted = cred_total_s1 + cred_total_s2

                deg_class = cal_class(yr_gp)
                            

                module_dir = os.path.dirname(__file__) 
                # get current directory
                if program.deg_id == "HND":
                    file_path1 = os.path.join(module_dir, 'templates/student_template/my_transcript_temp.html')
                    file_path2 = os.path.join(module_dir, 'templates/student_template/my_transcript.html')

                else:
                    file_path1 = os.path.join(module_dir, 'templates/student_template/my_transcript_btech_temp.html')
                    file_path2 = os.path.join(module_dir, 'templates/student_template/my_transcript_btech.html')
                    head = Image.objects.filter(title = 'trans_head_btech').first()
                    
                pwd = os.path.dirname(__file__)
                open(file_path1, "w").write(render_to_string(file_path2, {'data': sem1,
                                                                            'data1': sem2,
                                                                            'data2': resits,
                                                                            'regi':regi,
                                                                            'name':name,
                                                                            'phone': phone,
                                                                            'email':email,
                                                                            'cgpa':gpa_sem1,
                                                                            'status':grade_sem1,
                                                                            'gp1': gp_s1,
                                                                            'credt1': cred_total_s1,
                                                                            'wpt1':wpt_s1,
                                                                            'cgpa1':gpa_sem2,
                                                                            'status1':grade_sem2,
                                                                            'gp2': gp_s2,
                                                                            'credt2':cred_total_s2,
                                                                            'wpt2':wpt_s2,
                                                                            'cgpa2':gpa_resit,
                                                                            'gp3': gp_resit,
                                                                            'credt3':cred_total_resit,
                                                                            'wpt3':wpt_resit,
                                                                            'dept':dept,
                                                                            'ay': ay,
                                                                            'year_gpa': year_gpa,
                                                                            'year_grade': year_grade,
                                                                            'yr_gp': yr_gp,
                                                                            'deg_class': deg_class,
                                                                            'tcred': cred_attempted,
                                                                            'tcred_e': cred_earned,
                                                                            'program':program,
                                                                            'dob':dob,
                                                                            'pob':pob,
                                                                            'head': head,
                                                                            'level': level,
                                                                            'display': resit_display,
                                                                            }))

                # Converting the HTML template into a PDF file
                
                pdf = html_to_pdf(file_path1)

                pdfs.append(pdf)
                names.append(f_name)

        if len(pdfs) > 1:
            import zipfile
            output = io.BytesIO()
            with zipfile.ZipFile(output, 'w') as zf:
                for n,p in zip(names, pdfs):
                    zf.writestr(n,p)
                zf.close()

            response = HttpResponse(
            output.getvalue(),
                content_type="application/x-zip-compressed",
            )
            response["Content-Disposition"] = "attachment; filename=" + zip_name
                    
                    # rendering the template
            return response
        else:
            if len(pdfs) == 0:
                messages.success(request, "No Transcripts to print!")
                response = redirect('home')
            else:
                response = HttpResponse(
                pdfs[0],
                    content_type="application/pdf",
                )
                response["Content-Disposition"] = "attachment; filename=" + names[0]

            return response
    


##-----------------------------------------SEARCH------------------------------------####


@login_required(login_url = 'login')
@allowed_users(allowed_roles=['admin'])
def search_result1(request):
    #t_id = request.user.teacher.teacher_id
    data = Subject.objects.filter(semester = SemesterSession.objects.filter(active = "Yes").first().semester)

    context = {  'data':data }
    
    if request.method == 'POST':
        regi= request.POST.get('course_code')
        xx = regi.split(',')
        return redirect(reverse('search_result', kwargs={"course_code": xx[0], "dept" : xx[1]}, ))


    return render(request,'teacher_template/search_result1.html',context)





@login_required(login_url = 'login')
@allowed_users(allowed_roles=['teacher'])
def search_student_registered(request):

    t_id = str(request.user.teacher.teacher_id)
    data = AssignedTeacher2.objects.filter(teacher_id = t_id)
    
    context = {'data':data}
    
    if request.method == 'POST':
        course_id = request.POST.get('course_code')
        xx = course_id.split(",")

        return redirect(reverse('add_result', kwargs= {"dept": xx[1], "course_id": xx[0]}))
    return render(request,'teacher_template/search_student_registered.html',context)
        
###-----------------------------------SEARCH END------------------------------------------------------------------------#

@login_required(login_url = 'login')
@allowed_users(allowed_roles=['admin'])
def search_result(request, course_code, dept):
    data = Result.objects.filter(course_code = course_code)

    context={'course':data, 'course_code': course_code, 'dept': dept, 'resit': ['Yes', 'No']} 
    if request.method == 'POST':
        registration_number = request.POST.get('registration_number')
        resit = request.POST.get('resit')
        obj = Result.objects.get(student_id = registration_number ,course_code = course_code)
        id = int(obj.id)
        print(id)
        if resit == "Yes":
            return redirect(reverse('update_result_resit', kwargs= {"result_id": id,"course_code": course_code}))
        else:
            return redirect(reverse('update_result', kwargs= {"result_id": id,"course_code": course_code}))
        
    return render(request,'teacher_template/search_result.html',context)



@login_required(login_url = 'login')
@allowed_users(allowed_roles=['admin'])
def update_result(request, result_id, course_code):
    result = get_object_or_404(Result, id =result_id)
    stu = result.student_id
    course_id = result.course_code
    dept = result.dept
    theory_marks = result.theory_marks
    term_test = result.term_test
    total = result.total
    form = UpdateForm(request.POST or None, instance = result)
    regi = result.student

    context = {'form':form, 
            'regi': regi, 
            'course_id': course_id, 
    }
    
    if request.method == 'POST':
        if form.is_valid():
            print("form is valid")
            tt = form.instance.term_test
            theory= form.instance.theory_marks
            #form.instance.total = round((theory + tt )/5)
            form.save()
            result.save()
            messages.success(request,"Marks Edited for %s for %s "%(regi.name,course_id.subject_name))
            return redirect('search_result1')
    
    return render (request, 'teacher_template/update_result.html',context)

@login_required(login_url = 'login')
@allowed_users(allowed_roles=['admin'])
def update_result_resit(request, result_id, course_code):
    result = get_object_or_404(Result, id =result_id)
    stu = result.student_id
    course_id = result.course_code
    dept = result.dept
    theory_marks = result.theory_marks
    term_test = result.term_test
    total = result.total
    form = UpdateFormResit(request.POST or None, instance = result)
    regi = result.student

    context = {'form':form, 
            'regi': regi, 
            'course_id': course_id, 
    }
    
    if request.method == 'POST':
        result.resited = "Yes"
            

        test = request.POST.get('result')
        if regi.level == "HND1" or regi.level == "HND2" :
            result.term_test = test
        else:
            result.term_test_resit = int(test)
            

        result.save()
        
        messages.success(request,"Added Resit Result for %s for %s "%(regi.name,course_id.subject_name))
        return redirect('search_result1')
    
    return render (request, 'teacher_template/update_result_resit.html',context)


#----------------------------------------------** ADD ** -------------------------------------------------------------------------------------------------------


@login_required(login_url = 'login')
@allowed_users(allowed_roles=['admin'])
def add_student(request):
    user_form = CreateUserForm(request.POST or None) 
    student_form = StudentForm(request.POST or None, request.FILES or None)
    context = {'student_form': student_form,'user_form':user_form, 'page_title':'add student'}
    if request.method == 'POST':
        if user_form.is_valid and student_form.is_valid():
            user = user_form.save()
            student = student_form.save()
            student.user =user
            student.save()
            
            group = Group.objects.get(name = 'student')
            user.groups.add(group)
            messages.success(request, "Successfully Student Added")
        else:
            messages.error(request, "Form not Valid. Try another username and/or Registration Number and fill all fields correctly.")
    return render(request, 'student_template/add_student.html',context)


@login_required(login_url = 'login')
@allowed_users(allowed_roles=['teacher'])
def add_result(request, dept, course_id):

    register = RegisterTable.objects.filter(subject_id = course_id, dept_id = dept, status = 'Approved')
    context = {'data':register}
    if request.method == 'POST':
        regi = request.POST.get('registration_number')
        
        obj = Result.objects.filter(student_id = regi ,course_code = course_id).first()
        if obj != None:
            id = obj.id
            messages.info(request, "Result Already Exist, You Can Update That Result Here")
            return redirect(reverse('update_result', kwargs= {"result_id": id, "course_code": course_id}))
        else:
            return redirect(reverse('add_result2', kwargs= {"regi": regi, "cour_id": course_id}))

    return render(request, 'teacher_template/add_result.html',context)


@login_required(login_url = 'login')
@allowed_users(allowed_roles=['admin'])
def add_subject(request):
    subject_form = AddSubjectForm(request.POST or None, request.FILES or None)
    context = {'subject_form': subject_form, 'page_title':'add subject'}
    if request.method == 'POST':
        if subject_form.is_valid():
            subject = subject_form.save()
            subject.save()
            
            messages.success(request, "Successfully Added Subject")
        else:
            messages.error(request, "Form not Valid. Change course code and fill all fields correctly.")

    return render(request, 'admin_template/add_subject.html',context)



##Extracting result template
@login_required(login_url = 'login')
@allowed_users(allowed_roles=['teacher'])
def extract_temp(request):
    t_id = str(request.user.teacher.teacher_id)
    data1 = AssignedTeacher2.objects.filter(teacher_id = t_id) 

    c_ss = SemesterSession.objects.filter(active = 'Yes').first()
    disabled = ""
    data2 = []
    from datetime import datetime
    if datetime.now().date() > c_ss.ca_deadline and datetime.now().date() > c_ss.ca_deadline_btech and datetime.now().date() > c_ss.ca_deadline_masters:

        disabled = "disabled"
    if datetime.now().date() < c_ss.ca_deadline:
        for d in data1:
            if d.course.level == "HND1" or d.course.level == "HND2":
                data2.append(d)
    else:
        messages.success(request,"Deadline to Submit CA Results for HND has passed.")
    
    if datetime.now().date() < c_ss.ca_deadline_btech:
        for d in data1:
            if d.course.level == "BTECH":
                data2.append(d)
    else:
        messages.success(request,"Deadline to Submit CA Results for BTECH has passed.")
    
    if datetime.now().date() < c_ss.ca_deadline_masters:
        for d in data1:
            if d.course.level == "M1" or d.course.level == "M2":
                data2.append(d)
    else:
        messages.success(request,"Deadline to Submit CA Results for Masters has passed.")
    

    data = []
    for d in data2:
        if d.course.semester == c_ss.semester :
            data.append(d)

    
    context={'course':data, 'disabled': disabled} 
    if request.method == 'POST':
        code = request.POST.get('course_code')
        xx =code.split(",")
        course_cd = xx[0]
        dept_id =xx[1]

        matricules = []
        names = []
        fields = ["Matricule", "Names", "CA"]

        register1 = RegisterTable.objects.filter(subject_id = course_cd, sem_ses = c_ss).all()
        for r in register1:
            matricules.append(r.student)

            n = Student.objects.filter(registration_number = r.student).first().name
            names.append(n)

        course_name = Subject.objects.filter(course_code = course_cd).first().subject_name
        wb_name = 'CA Results_' + str(course_name) + '.xlsx'

        output = io.BytesIO()
        wb = xlsxwriter.Workbook(output)
        ws = wb.add_worksheet('Sheet1')
        
        r1 = 0
        c1 = 0

        for f in fields:
            ws.write(r1, c1, f)
            c1+=1
        r2 = 1
        c2 = 0

        import itertools
        for (m, n) in zip(matricules, names):
            ws.write(r2, c2, str(m))
            ws.write(r2,c2 + 1, str(n))
            r2+=1

        wb.close()
        
        output.seek(0)

        response = HttpResponse(
            output,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = "attachment; filename=%s" % wb_name

        return response
            

    return render(request,'teacher_template/extract_temp.html',context)

@login_required(login_url = 'login')
@allowed_users(allowed_roles=['teacher'])
def extract_temp_exam(request):
    t_id = str(request.user.teacher.teacher_id)
    data1 = AssignedTeacher2.objects.filter(teacher_id = t_id)

    c_ss = SemesterSession.objects.filter(active = 'Yes').first()
    disabled = ""
    data2 = []
    from datetime import datetime
    if datetime.now().date() > c_ss.result_deadline and datetime.now().date() > c_ss.result_deadline_btech and datetime.now().date() > c_ss.result_deadline_masters:

        disabled = "disabled"
    if datetime.now().date() < c_ss.result_deadline:
        for d in data1:
            if d.course.level == "HND1" or d.course.level == "HND2":
                data2.append(d)
    else:
        messages.success(request,"Deadline to Submit Exam Results for HND has passed.")
    
    if datetime.now().date() < c_ss.result_deadline_btech:
        for d in data1:
            if d.course.level == "BTECH":
                data2.append(d)
    else:
        messages.success(request,"Deadline to Submit Exam Results for BTECH has passed.")
    
    if datetime.now().date() < c_ss.result_deadline_masters:
        for d in data1:
            if d.course.level == "M1" or d.course.level == "M2":
                data2.append(d)
    else:
        messages.success(request,"Deadline to Submit Exam Results for Masters has passed.")
    

    data = []
    for d in data2:
        if d.course.semester == c_ss.semester :
            data.append(d)

    context={'course':data, 'disabled': disabled} 
    if request.method == 'POST':
        code = request.POST.get('course_code')
        xx =code.split(",")
        course_cd = xx[0]
        dept_id =xx[1]

        codes = []
        fields = ["Code", "Exam Mark"]

        register1 = RegisterTable.objects.filter(subject_id = course_cd, sem_ses = c_ss).all()
        for r in register1:
            c = ExamCode.objects.filter(student = r.student, sem_ses = c_ss, subject = r.subject).first().code
            codes.append(c)

        course_name = Subject.objects.filter(course_code = course_cd).first().subject_name
        wb_name = 'Exam Results_' + str(course_name) + '_' + str(c_ss.ss_id) + '.xlsx'

        output = io.BytesIO()
        wb = xlsxwriter.Workbook(output)
        ws = wb.add_worksheet('Sheet1')
        
        r1 = 0
        c1 = 0

        for f in fields:
            ws.write(r1, c1, f)
            c1+=1
        r2 = 1
        c2 = 0

        
        for c in codes:
            ws.write(r2, c2, str(c))
            r2+=1

        wb.close()
        
        output.seek(0)

        response = HttpResponse(
            output,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = "attachment; filename=%s" % wb_name

        return response
            

    return render(request,'teacher_template/extract_temp_exam.html',context)

@login_required(login_url = 'login')
@allowed_users(allowed_roles=['admin'])
def extract_res_stat(request):

    ss = SemesterSession.objects.all()
    depts = Dept.objects.all()
    levs = ['HND1','HND2','BTECH','M1','M2']


    context={'sems':ss, 'depts': depts, 'levs': levs} 
    if request.method == 'POST':
        sem = request.POST.get('sem')
        dept = request.POST.get('dept')
        lev = request.POST.get('lev')

        res = Result.objects.filter(sem_ses = sem, level = lev, dept = dept).all()

        if not res:
            messages.success(request, "No results.")
        else:
            fields = ['Course Code', 'Course Title', 'Credit Value', 'Course Instructor', 'Registered', 'Examined', 'Passed', 'Failed', '% Passed','A','B+','B','C+','C','D+','D','F']

            codes = []
            titles = [] 
            creditvs = [] 
            instructors = [] 
            reg = []
            examd = []
            passed = []
            failed = []
            per_passed = []
            As = []
            Bps = []
            Bs = []
            Cps = []
            Cs = []
            Dps = []
            Ds = []
            Fs = []

            for r in res:
                if r.course_code.course_code in codes:
                    continue
                else:
                    codes.append(r.course_code.course_code)
                    titles.append(r.course_code.subject_name)
                    creditvs.append(r.course_code.credit)
                    instructors.append(AssignedTeacher2.objects.filter(dept = dept, course = r.course_code).first().teacher)
                    reg.append(RegisterTable.objects.filter(sem_ses = sem, dept = dept, subject = r.course_code).all().count())
                    ex = res.filter(course_code = r.course_code).all()
                    A = 0
                    Bp = 0
                    B = 0
                    Cp = 0
                    C = 0
                    Dp = 0
                    D = 0
                    F = 0
                    for r in ex:
                        grade = cal_cgname(cal_cg(r.total))
                        if grade == 'A':
                            A+=1
                        elif grade == 'B+':
                            Bp+=1
                        elif grade == 'B':
                            B+=1
                        elif grade == 'C+':
                            Cp+=1
                        elif grade == 'C':
                            C+=1
                        elif grade == 'D+':
                            Dp+=1
                        elif grade == 'D':
                            D+=1
                        else:
                            F+=1
                    As.append(A)
                    Bps.append(Bp)
                    Bs.append(B)
                    Cps.append(Cp)
                    Cs.append(C)
                    Dps.append(Dp)
                    Ds.append(D)
                    Fs.append(F)

                    examd.append(ex.count())
                    psd = res.filter(course_code = r.course_code, total__gte = 10).all().count()
                    passed.append(psd)
                    failed.append(res.filter(course_code = r.course_code, total__lt = 10).all().count())
                    per_passed.append(round((psd / ex.count() )*100,2))
                

            wb_name = 'Results_Stats_' + str(dept) + '_' + str(lev) + '_' + str(sem) + '.xlsx'

            output = io.BytesIO()
            wb = xlsxwriter.Workbook(output)
            ws = wb.add_worksheet('Sheet1')

            ws.write(0,0, 'Results Statistics ' + dept + ' for ' + sem + '')
            
            r1 = 1
            c1 = 0

            for f in fields:
                ws.write(r1, c1, f)
                c1+=1

            r2 = 2
            c2 = 0
            for (c,t,cv,i,r,e,p,fd,pp,a,bp,b,cp,ce,dp,d,ef) in zip(codes, titles, creditvs, instructors, reg, examd, passed, failed, per_passed,As,Bps,Bs,Cps,Cs,Dps,Ds,Fs):
                ws.write(r2, c2, str(c))
                ws.write(r2, c2 + 1, str(t))
                ws.write(r2, c2 + 2, cv)
                ws.write(r2, c2 + 3, str(i))
                ws.write(r2, c2 + 4, r)
                ws.write(r2, c2 + 5, e)
                ws.write(r2, c2 + 6, p)
                ws.write(r2, c2 + 7, fd)
                ws.write(r2, c2 + 8, pp)
                ws.write(r2, c2 + 9, a)
                ws.write(r2, c2 + 10, bp)
                ws.write(r2, c2 + 11, b)
                ws.write(r2, c2 + 12, cp)
                ws.write(r2, c2 + 13, ce)
                ws.write(r2, c2 + 14, dp)
                ws.write(r2, c2 + 15, d)
                ws.write(r2, c2 + 16,ef)
                
                r2+=1

            r2+=2
            ws.write(r2, c2 + 1, "Total")
            ws.write(r2, c2 + 4, sum(reg))
            exd = sum(examd)
            ws.write(r2, c2 + 5, exd)
            psd = sum(passed)
            ws.write(r2, c2 + 6, psd)
            ws.write(r2, c2 + 7, sum(failed))
            ws.write(r2, c2 + 8, round((psd / exd)*100,2))

            wb.close()
            
            output.seek(0)

            response = HttpResponse(
                output,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = "attachment; filename=%s" % wb_name

            return response
            

    return render(request,'admin_template/result_stats.html',context)

@login_required(login_url = 'login')
@allowed_users(allowed_roles=['teacher'])
def add_excel(request):

    t_id = str(request.user.teacher.teacher_id)
    data1 = AssignedTeacher2.objects.filter(teacher_id = t_id)

    c_ss = SemesterSession.objects.filter(active = 'Yes').first()
    disabled = ""
    data2 = []
    from datetime import datetime
    if datetime.now().date() > c_ss.ca_deadline and datetime.now().date() > c_ss.ca_deadline_btech and datetime.now().date() > c_ss.ca_deadline_masters:

        disabled = "disabled"
    if datetime.now().date() < c_ss.ca_deadline:
        for d in data1:
            if d.course.level == "HND1" or d.course.level == "HND2":
                data2.append(d)
    else:
        messages.success(request,"Deadline to Submit CA Results for HND has passed.")
    
    if datetime.now().date() < c_ss.ca_deadline_btech:
        for d in data1:
            if d.course.level == "BTECH":
                data2.append(d)
    else:
        messages.success(request,"Deadline to Submit CA Results for BTECH has passed.")
    
    if datetime.now().date() < c_ss.ca_deadline_masters:
        for d in data1:
            if d.course.level == "M1" or d.course.level == "M2":
                data2.append(d)
    else:
        messages.success(request,"Deadline to Submit CA Results for Masters has passed.")
    
    data = []
    for d in data2:
        if d.course.semester == c_ss.semester :
            data.append(d)

    context={'course':data, 'disabled': disabled} 
    if request.method == 'POST' and request.FILES['myfile']:
        myfile = request.FILES['myfile']
        code = request.POST.get('course_code')
        xx =code.split(",")
        course_cd = xx[0]
        dept_id =xx[1]

        wb = openpyxl.load_workbook(myfile)
        ws = wb["Sheet1"]

        excel_data = []

        for r in ws.iter_rows():
            row_data = []
            for cell in r:
                if cell.value is None:
                    row_data.append("0")
                else:   
                    row_data.append(str(cell.value))
            excel_data.append(row_data)
        
        success = False
        for i in range(1, len(excel_data)):
            try:
                stud = Student.objects.filter(registration_number = excel_data[i][0]).first()
                res = Result.objects.filter(sem_ses = c_ss, student = stud, course_code = course_cd).first()  

                if res != None:
                    res.theory_marks = excel_data[i][2]
                    res.save()
                    success = True
                else:
                    sub = Result(
                            sem_ses = c_ss,
                            course_code =Subject.objects.filter(course_code = course_cd).first(),
                            theory_marks = excel_data[i][2],
                            dept = dept_id,
                            student_id = excel_data[i][0],
                            level = Student.objects.filter(registration_number = excel_data[i][0]).first().level
                            )
                    sub.save()
                    success = True
            except IntegrityError :
                messages.success(request,"Student %s already has Result for %s course"% (excel_data[i][0], course_cd))
                success = False
                continue
            except DataError:
                messages.success(request, "An error occured.You may have used wrong file. Check your file and make sure you are using correct file. If error persists contact platform admin.")
                success = False
            except ValueError:
                messages.success(request, "An error occured.You may have used wrong file. Check your file and make sure you are using correct file. If error persists contact platform admin.")
                success = False
        
        if success:
            messages.success(request,"Successfully Added Results for %s course"% (course_cd))
        return redirect('home') 
    return render(request,'teacher_template/add_json.html',context)

@login_required(login_url = 'login')
@allowed_users(allowed_roles=['teacher'])
def add_exam(request):
    t_id = str(request.user.teacher.teacher_id)
    data1 = AssignedTeacher2.objects.filter(teacher_id = t_id)

    c_ss = SemesterSession.objects.filter(active = 'Yes').first()
    disabled = ""
    data2 = []
    from datetime import datetime
    if datetime.now().date() > c_ss.result_deadline and datetime.now().date() > c_ss.result_deadline_btech and datetime.now().date() > c_ss.result_deadline_masters:

        disabled = "disabled"
    if datetime.now().date() < c_ss.result_deadline:
        for d in data1:
            if d.course.level == "HND1" or d.course.level == "HND2":
                data2.append(d)
    else:
        messages.success(request,"Deadline to Submit Exam Results for HND has passed.")
    
    if datetime.now().date() < c_ss.result_deadline_btech:
        for d in data1:
            if d.course.level == "BTECH":
                data2.append(d)
    else:
        messages.success(request,"Deadline to Submit Exam Results for BTECH has passed.")
    
    if datetime.now().date() < c_ss.result_deadline_masters:
        for d in data1:
            if d.course.level == "M1" or d.course.level == "M2":
                data2.append(d)
    else:
        messages.success(request,"Deadline to Submit Exam Results for Masters has passed.")
    

    data = []
    for d in data2:
        if d.course.semester == c_ss.semester :
            data.append(d)

    context={'course':data, 'disabled': disabled} 
    if request.method == 'POST' and request.FILES['myfile']:
        myfile = request.FILES['myfile']
        code = request.POST.get('course_code')
        xx =code.split(",")
        course_cd = xx[0]
        dept_id =xx[1]

        wb = openpyxl.load_workbook(myfile)
        ws = wb["Sheet1"]

        excel_data = []

        for r in ws.iter_rows():
            row_data = []
            for cell in r:
                if cell.value is None:
                    row_data.append("0")
                else:   
                    row_data.append(str(cell.value))
            excel_data.append(row_data)
        
        success = False
        for i in range(1, len(excel_data)):
            try:
                stud = ExamCode.objects.filter(code = excel_data[i][0], sem_ses = c_ss, subject = course_cd).first()

                res = Result.objects.filter(sem_ses = c_ss, student = stud.student, course_code = course_cd).first()
                res.term_test = excel_data[i][1]
                res.save()
                success = True
            except AttributeError as e :
                messages.success(request,"The following error occured %s You may have uploaded wrong file. Make sure you have uploaded the appropriate file and try again. If error persists see platform administrator."% (e))
                break
            except ValueError as e :
                messages.success(request,"The following error occured %s You may have uploaded wrong file. Make sure you have uploaded the appropriate file and try again. If error persists see platform administrator."% (e))
                break
        if success:
            messages.success(request,"Successfully Added Results for %s course"% (course_cd))
        return redirect('home') 
    return render(request,'teacher_template/add_exam.html',context)


@login_required(login_url = 'login')
@allowed_users(allowed_roles=['admin'])         
def add_students(request):
    data = Dept.objects.all()
    context={'course':data} 
    if request.method == 'POST' and request.FILES['myfile']:
        myfile = request.FILES['myfile']
        dept_id = request.POST.get('dept')
        dept = Dept.objects.filter(dept_id = dept_id).first()


        wb = openpyxl.load_workbook(myfile)
        ws = wb["Sheet1"]

        excel_data = []

        for r in ws.iter_rows():
            row_data = []
            for cell in r:
                if cell.value is None:
                    row_data.append("0")
                else:   
                    row_data.append(str(cell.value))
            excel_data.append(row_data)
        
        
        for i in range(1, len(excel_data)):
            try:
                user = User(
                    username = excel_data[i][0],
                    email = excel_data[i][1],
                )
                user.set_password(excel_data[i][2])
                user.save()
                
                import pandas as pd
                stud = Student(
                        user = user,
                        registration_number = excel_data[i][5],
                        dept = dept,
                        name = excel_data[i][3],
                        phone = excel_data[i][4],
                        level = excel_data[i][8],
                        dob = pd.to_datetime(excel_data[i][6]).date(),
                        pob = excel_data[i][7],
                        degree_pursued =  Degree.objects.filter(deg_id = excel_data[i][9]).first()
                    )
                stud.save()
                group = Group.objects.get(name = 'student')
                user.groups.add(group)
                messages.success(request,"Successfully Added Student %s for %s Department"% (excel_data[i][3],dept_id))
            except IntegrityError:
                messages.success(request,"User with given username already exists or Student with given registration number already exists")
                continue
            except DataError:
                messages.success(request, "An error occured.You may have used wrong file. Check your file and make sure you are using correct file. If error persists contact platform admin.")
                return redirect('home')
            except ValueError:
                messages.success(request, "An error occured.You may have used wrong file. Check your file and make sure you are using correct file. If error persists contact platform admin.")
                return redirect('home')

        
        return redirect('home') 
    return render(request,'student_template/add_students.html',context)  

@login_required(login_url = 'login')
@allowed_users(allowed_roles=['admin'])         
def reupload_results(request):
    ss = SemesterSession.objects.filter(active ='Yes').first()
    depts = Dept.objects.all()
    levs = ['HND1','HND2','BTECH','M1','M2']


    context={ 'depts': depts, 'levs': levs} 

    if request.method == 'POST' and request.FILES['myfile']:
        myfile = request.FILES['myfile']
        dept_id = request.POST.get('dept')
        dept = Dept.objects.filter(dept_id = dept_id).first()
        level = request.POST.get('level')

        try:
            wb = openpyxl.load_workbook(myfile)
            ws = wb["Sheet1"]

            excel_data = []

            for r in ws.iter_rows():
                row_data = []
                for cell in r:   
                    if cell.value is None:
                        row_data.append("0")
                    else:   
                        row_data.append(str(cell.value))
                excel_data.append(row_data)
            

            subjects = Subject.objects.filter(dept = dept_id, level = level, semester = ss.semester).all()


            for i in range(1, len(excel_data)):
                j=4
                while j < subjects.count() + 4:
                    for s in subjects:  
                        r = Result.objects.filter(sem_ses = ss, student = excel_data[i][0], level = level, course_code = s ).first()
                        r.term_test = excel_data[i][j]
                        r.save()
                        j+=1
            messages.success(request, "Succesfully reuploaded results.")
            return redirect('home') 
        except OSError:
            messages.success(request, "An error occured.Please Upload an excel file. If error persists contact platform admin.")
            return redirect('home')
        except AttributeError:
            messages.success(request, "Attribute Errot. Params: Student: {stud}, sem_ses: {sem_ses}, level: {level}, course_code: {course_code}".format(stud = excel_data[i][0], sem_ses = ss.ss_id, level = level, course_code = s.course_code ))
            return redirect('home')
        except ValueError:
            messages.success(request, "An error occured.You may have used wrong file. Check your file and make sure you are using correct file. If error persists contact platform admin.")
            return redirect('home')
                
    return render(request,'admin_template/reupload_results.html',context)  

@login_required(login_url = 'login')
@allowed_users(allowed_roles=['admin'])         
def reupload_results_ca(request):
    ss = SemesterSession.objects.filter(active ='Yes').first()
    depts = Dept.objects.all()
    levs = ['HND1','HND2','BTECH','M1','M2']


    context={ 'depts': depts, 'levs': levs} 

    if request.method == 'POST' and request.FILES['myfile']:
        myfile = request.FILES['myfile']
        dept_id = request.POST.get('dept')
        dept = Dept.objects.filter(dept_id = dept_id).first()
        level = request.POST.get('level')

        try:
            wb = openpyxl.load_workbook(myfile)
            ws = wb["Sheet1"]

            excel_data = []

            for r in ws.iter_rows():
                row_data = []
                for cell in r:
                    if cell.value is None:
                        row_data.append("0")
                    else:   
                        row_data.append(str(cell.value))
                excel_data.append(row_data)

            subjects = Subject.objects.filter(dept = dept_id, level = level, semester = ss.semester).all()

            for i in range(1, len(excel_data)):
                j=4
                while j < subjects.count() + 4:
                    for s in subjects:
                        r = Result.objects.filter(sem_ses = ss, student = excel_data[i][0], level = level, course_code = s ).first()
                        r.theory_marks = excel_data[i][j]
                        r.save()
                        j+=1
            messages.success(request, "Succesfully reuploaded results.")
            return redirect('home') 
        except OSError:
            messages.success(request, "An error occured.Please Upload an excel file. If error persists contact platform admin.")
            return redirect('home')
        except AttributeError:
            messages.success(request, " Attribute Errot. Params: Student: {stud}, sem_ses: {sem_ses}, level: {level}, course_code: {course_code}".format(stud = excel_data[i][0], sem_ses = ss.ss_id, level = level, course_code = s.course_code ))
            return redirect('home')
        except ValueError:
            messages.success(request, "An error occured.You may have used wrong file. Check your file and make sure you are using correct file. If error persists contact platform admin.")
            return redirect('home')
                
    return render(request,'admin_template/reupload_results_ca.html',context)  

@login_required(login_url = 'login')
@allowed_users(allowed_roles=['admin'])
def add_admin(request):
    return redirect('register')

@login_required(login_url = 'login')
@allowed_users(allowed_roles=['admin'])
def addDept(request):
    form = DepartmentForm(request.POST or None)
    context = {'form': form, 'page_title':'add department'}
    if request.method == 'POST':
        if form.is_valid():
            try:
                form.save()
                messages.success(request,"Successfully Added Dept. ")
                return redirect('home')
            except IntegrityError:
                messages.success(request,"Department with filled Dept ID already exists. Try another ID please.")
                return redirect('home')

    return render(request, 'registration_template/add_dept.html',context)

@login_required(login_url = 'login')
@allowed_users(allowed_roles=['admin'])
def addSchool(request):
    form = SchoolForm(request.POST or None)
    context = {'form': form, 'page_title':'add school'}
    if request.method == 'POST':
        if form.is_valid:
            try:
                form.save()
                messages.success(request,"Successfully Added School")
            except IntegrityError:
                messages.success(request,"School with filled School ID already exists. Try another School ID please.")

    return render(request, 'registration_template/add_school.html',context)

@login_required(login_url = 'login')
@allowed_users(allowed_roles=['admin'])
def addDegree(request):
    form = DegreeForm(request.POST or None)
    context = {'form': form, 'page_title':'add degree'}
    if request.method == 'POST':
        if form.is_valid:
            try:
                form.save()
                messages.success(request,"Successfully Added Degree")
            except IntegrityError:
                messages.success(request,"Degree with filled Degree ID already exists. Try another Degree ID please.")


    return render(request, 'registration_template/add_deg.html',context)

@login_required(login_url = 'login')
@allowed_users(allowed_roles=['admin'])
def addSS(request):
    form = SSForm(request.POST or None)
    context = {'form': form, 'page_title':'add semester session'}
    if request.method == 'POST':
        if form.is_valid:
            try:
                form.save()
                messages.success(request,"Successfully Added Semester Session")
            except IntegrityError:
                messages.success(request,"Semester Session with filled ID already exists. Try another ID please.")


    return render(request, 'registration_template/add_ss.html',context)

@login_required(login_url = 'login')
@allowed_users(allowed_roles=['admin'])
def setActiveSS(request):
    data = SemesterSession.objects.all()

    active = SemesterSession.objects.filter(active = 'Yes').first()

    context = {'ss' : data, 'current': active}

    if request.method == 'POST':
        sems = SemesterSession.objects.all()
        for s in sems:
            s.active = 'No'
            s.save()
        
        ssid = request.POST.get('ssid')

        sem = SemesterSession.objects.filter(ss_id = ssid).first()
        sem.active = 'Yes'
        sem.save()

        ''' Register all students to various courses for the semester
        students = Student.objects.all()
        print(students)
        for stud in students:
            print(stud)
            deg = stud.degree_pursued.deg_id
            lev = stud.level
            dep = stud.dept

            if deg == "HND":
                sub = "HND"
            elif deg == "BTECH":
                sub = "DEGREE"
            else:
                sub = "MASTER"

            courses = Subject.objects.filter(level = lev, subtype = sub, dept= dep, semester = sem.semester)

            for c in courses:
                print(c)
                check = RegisterTable.objects.filter(student = stud, subject = c, sem_ses = sem).first()
                if check == None:
                    print('registering student: ' + str(stud) + ' to course: ' + str(c.subject_name))
                    ss = RegisterTable(
                        sem_ses = sem,
                        dept = dep,
                        student = stud,
                        subject = c
                    )
                    ss.save()

                    tid = AssignedTeacher2.objects.filter(course_id = c.course_code).first().teacher.teacher_id

                    print(tid)

                    rr = Rating.objects.filter(student = stud , subject_id = c.course_code, teacher_id = tid).first()
                    if rr == None:
                        rate = Rating(
                            student = stud,
                            subject_id = c.course_code,
                            teacher_id = tid,
                        )
                        rate.save()
                    else:
                        print('Continuing loop because rating exist...')
                        continue
                    
                else:
                    print('Continuing loop because student already registered...')
                    continue
            '''
        messages.success(request,"Semester:  " + sem.ss_id + " is set active.") 

        return redirect('home')
    
    return render(request,'admin_template/set_active_ss.html', context)

@login_required(login_url = 'login')
@allowed_users(allowed_roles=['admin'])
def assign_teacher_dept_search(request):
    #data = Result.objects.raw('''
    #    SELECT dept_id as id FROM app_dept''')
    data = Dept.objects.all()

    context={'dept':data} 
    if request.method == 'POST':
        dept_id = request.POST.get('dept_id')
        return redirect(reverse('assign_teacher', kwargs={"dept_id": dept_id}))

    return render(request,'admin_template/assign_teacher_dept_search.html',context)

def assign_teacher(request, dept_id):
    data = Teacher.objects.all()
    course = Subject.objects.filter(dept = dept_id)
    context = {'teacher':data, 'course': course, 'teacher_dept':dept_id}

    if request.method == "POST":

        cour_code = request.POST.get('course')
        t_id   = request.POST.get('teacher')
        t_dept = dept_id

        exists = AssignedTeacher2.objects.filter(course = Subject.objects.filter(course_code = cour_code).first(), teacher_id = t_id).first()
        exists2 = AssignedTeacher2.objects.filter(course = Subject.objects.filter(course_code = cour_code).first())
        
        if exists:
            messages.success(request,"This Lecturer is already assigned to this course")
        elif exists2:
            messages.success(request,"Lecturer %s is already assigned to this course. You cannot assign 2 lecturers to same course. See platform admin to remove previous assignment."%(exists2.first().teacher.name))
        else:

            assingn = AssignedTeacher2(
                course = Subject.objects.filter(course_code = cour_code).first(),
                dept_id = t_dept,
                
                teacher_id= t_id,

            )
            assingn.save()
            messages.success(request,"Lecturer Id : %s Is Assigned For %s Course In %s Department" %(t_id,cour_code,t_dept))

    return render(request,'admin_template/assign_teacher_dept.html',context)

@login_required(login_url = 'login')
@allowed_users(allowed_roles=['admin'])
def change_stud_dept(request):
    data1 = Student.objects.all()
    
    data2 = Dept.objects.all()
    


    context = {'studs' : data1, 'depts': data2}

    if request.method == 'POST':
        stud = request.POST.get('stud')
        dept = request.POST.get('dept')

        d = Dept.objects.filter(dept_id = dept).first()
        Student.objects.filter(registration_number = stud).update(dept = d)
        #student.dept = d
        #student.save()
        messages.success(request,"Student:  " + stud + " transfered to " + dept + " Department")
    
    return render(request,'admin_template/change_stud_dept.html', context)

@login_required(login_url = 'login')
@allowed_users(allowed_roles=['admin'])
def promote_stud(request):
    data1 = Dept.objects.all()
    
    data2 = ['HND1','HND2','BTECH','M1','M2']


    context = {'depts' : data1, 'levs': data2}

    if request.method == 'POST':
        dept = request.POST.get('dept')
        lev = request.POST.get('lev')

        return redirect(reverse('promote_student2', kwargs={"dept": dept, "lev" : lev}, ))

        
        messages.success(request,"Student:  " + stud + " promoted to " + lev + " Level")
    
    return render(request,'admin_template/promote_stud.html', context)

@login_required(login_url = 'login')
@allowed_users(allowed_roles=['admin'])
def promote_stud2(request, dept, lev):
    studs = Student.objects.filter(dept_id = dept, level = lev).all()
    data2 = ['HND1','HND2','BTECH','M1','M2']

    context = {"studs": studs, 'levs': data2, 'lev': lev, 'dept': dept}

    if request.method == 'POST':
        lev = request.POST.get('lev')
        studs = request.POST.getlist('studs')
        if lev == 'BTECH':
            for stud in studs:
                deg = Degree.objects.filter(deg_id = lev).first()
                Student.objects.filter(registration_number = stud).update(degree_pursued = deg)
                Student.objects.filter(registration_number = stud).update(level = lev)
        elif lev == 'M1' or lev == 'M2':
            for stud in studs:
                deg = Degree.objects.filter(deg_id = 'MSC').first()
                Student.objects.filter(registration_number = stud).update(degree_pursued = deg)
                Student.objects.filter(registration_number = stud).update(level = lev)
        else :
            for stud in studs:
                deg = Degree.objects.filter(deg_id = "HND").first()
                Student.objects.filter(registration_number = stud).update(degree_pursued = deg)
                Student.objects.filter(registration_number = stud).update(level = lev)
            

    return render(request, 'admin_template/promote_stud2.html', context)

@login_required(login_url = 'login')
@allowed_users(allowed_roles=['admin'])
def block_stud(request):
    data1 = Student.objects.all()
    
    context = {'studs' : data1}

    if request.method == 'POST':
        stud = request.POST.get('stud')
        
        student = Student.objects.filter(registration_number = stud).first()
        student.user.is_active = False 
        student.user.save()
        student.save()

        messages.success(request,"Student:  " + student.name + " blocked. The student can no more access his account. ")
    
    return render(request,'admin_template/block_stud.html', context)

@login_required(login_url = 'login')
@allowed_users(allowed_roles=['admin'])
def unblock_stud(request):
    data1 = Student.objects.all()
    studs = []
    for s in data1:
        if s.user.is_active == False:
            studs.append(s)
    
    context = {'studs' : studs}

    if request.method == 'POST':
        stud = request.POST.get('stud')
        
        student = Student.objects.filter(registration_number = stud).first()
        student.user.is_active = True
        student.user.save()
        student.save()

        messages.success(request,"Student:  " + student.name + " unblocked. The student can now access his account again. ")

    return render(request,'admin_template/unblock_stud.html', context)



#--------------------------------------------------------###### ADD END #######---------------------------------------------------------



def student_sub_register(request):
    dept_name = request.user.student.dept
    deg = request.user.student.degree_pursued.deg_id
    regi = request.user.student
    sem = SemesterSession.objects.filter(active = 'Yes').first()

    if deg == "HND":
        sub = "HND"
    elif deg == "BTECH":
        sub = "DEGREE"
    else:
        sub = "MASTER"

    lev = str(request.user.student.level)

    data = Subject.objects.filter(dept= dept_name, subtype = sub)

    for i in data:
        ctt = RegisterTable.objects.filter(subject_id = i.course_code, student = regi, sem_ses = sem).first()
        
        if ctt != None:
            i.subject_name = i.subject_name + "-->Already Registered."
            
    context = {'data':data, 'aca': sem.session}

    if request.method == "POST":
        course_cc = request.POST.get('course_regi')
        regi = request.user.student

        check = RegisterTable.objects.filter(student = regi, subject_id = course_cc, sem_ses = sem).first()
        if check == None:
            
            ss = RegisterTable(
                sem_ses = sem,
                dept = dept_name,
                student = regi,
                subject_id = course_cc
            )
            ss.save()

            tid = AssignedTeacher2.objects.filter(course_id = course_cc).first().teacher.teacher_id


            rr = Rating.objects.filter(student = regi , subject_id = course_cc, teacher_id = tid).first()
            if rr == None:
                rate = Rating(
                    student = regi,
                    subject_id = course_cc,
                    teacher_id = tid,
                )
                rate.save()
            messages.success(request, "Registration Successful")
            
        else:
            messages.success(request, "You have already Registered for  this Subject.")   

        return redirect('home')
    return render(request,'teacher_template/student_sub_register.html',context)


def teacher_approve_search(request):
    teach_id = request.user.teacher.teacher_id
    data = AssignedTeacher2.objects.filter(teacher_id = teach_id)

    context= {'data':data}
    if request.method == "POST":
        course_code_dept = request.POST.get('course_code_dept')
        xx = course_code_dept.split(',')
        return redirect(reverse('teacher_approval', kwargs= {"course_code": xx[0], "student_dept": xx[1]}))

    return render(request, 'teacher_template/teacher_approve_search.html',context)


def teacher_approval(request, course_code, student_dept):
    data2 = Result.objects.raw('''
        SELECT 1 as id, name, student_id, phone, status, profile_pic  FROM
        public.app_student JOIN public.app_registertable ON
        app_student.registration_number = app_registertable.student_id
	    where app_registertable.dept_id = %s and app_registertable.subject_id = %s''',[student_dept,course_code]) 
    data = RegisterTable.objects.filter(subject_id = course_code, dept_id = student_dept)
    stu = []
    for i in data:
        stu.append( Student.objects.filter(registration_number = i.student_id).first())

    
        
    context ={'data':data2,'stu':stu, 'cc': course_code, 'dpt': student_dept}
    if request.method == "POST":
        stat = request.POST.get('optionsRadios')
        xx = stat.split(',')
        if xx[0]=='Pending' or xx[0]=='Rejected':
            Result.objects.filter(student_id = xx[1], course_code = course_code, dept = student_dept).delete()

        RegisterTable.objects.filter(subject_id = course_code, student_id = xx[1]).update(status = xx[0])
        
        tid = str(request.user.teacher.teacher_id)

        if xx[0] == 'Approved':
            rr = Rating.objects.filter(student_id = xx[1], subject_id = course_code, teacher_id = tid).first()
            if rr == None:
                rate = Rating(
                    student_id = xx[1],
                    subject_id = course_code,
                    teacher_id = tid,
                )
                rate.save()

    return render(request, 'teacher_template/teacher_approval.html',context)


def student_rating(request):
    c_ss = SemesterSession.objects.filter(active = 'Yes').first()
    regi = request.user.student

    data = RegisterTable.objects.filter(student = regi, sem_ses = c_ss).all()
    teachers = []
    for d in data:
        teachers.append(AssignedTeacher2.objects.filter(course = d.subject).first())
    

    if request.method == "POST":
        stat = request.POST.get('optionsRadios')
        xx = stat.split(',')
    
        Rating.objects.filter(subject_id = xx[2], student = regi,teacher_id = xx[1]).update(rating = int(xx[0]))

        return redirect('home')

    context = { 'regi':regi,
                'data':teachers,
                'aca': c_ss.session,
    }

    return render(request, 'student_template/student_rating.html',context)


def get_ratings_teacher(request):
    t_id = str(request.user.teacher.teacher_id)
    attendance = Result.objects.raw('''
    SELECT 1 as id, subject_id as sn , AVG(rating) as avg FROM
    public.app_rating
	where teacher_id=%s
	group by subject_id;''',[t_id])
    data =[]
    labels =[]

    for i in attendance:
        labels.append(i.sn)
        data.append(round(i.avg,2))

    return JsonResponse(data={
        'labels': labels,
        'data':data,
    })

def get_ratings_admin(request):
    t_rate = Result.objects.raw('''
    SELECT 1 as id, app_rating.teacher_id as tid ,app_rating.teacher_id as nem , AVG(rating) as avg FROM
    public.app_rating JOIN app_teacher ON app_rating.teacher_id = app_teacher.teacher_id
	group by app_rating.teacher_id ORDER BY AVG(rating) DESC;''')

    data =[]
    labels =[]

    for i in t_rate:
        nem = Teacher.objects.get(teacher_id = i.tid).name
        dep = str(Teacher.objects.get(teacher_id = i.tid).dept_id)
        i.nem = nem
        labels.append(i.nem + " ["+ i.tid+" ("+ dep + ") "+"]")
        data.append(round(i.avg,2))
    
    return JsonResponse(data={
        'labels': labels,
        'data':data,
    })

def dept_performance(request):
    dep_per = Result.objects.raw('''
    SELECT 1 as id, AVG(total) as avg, dept FROM
    public.app_result 
	group by dept ORDER BY AVG(total) DESC;''')
    data =[]
    labels =[]
    for i in dep_per:
        data.append((i.avg/20)*10)
        labels.append(i.dept)
    return JsonResponse(data={
        'labels': labels,
        'data':data,
    })


@login_required(login_url = 'login')
def subject_ranksheet_teacher(request):
    t_id = request.user.teacher.teacher_id
        
    data = AssignedTeacher2.objects.filter(teacher_id = t_id)
    data2 = []

    for d in data:
        if d.course.semester == SemesterSession.objects.filter(active = 'Yes').first().semester:
            data2.append(d)
    
    if len(data2) == 0:
        messages.success(request,"Seems you have no course assigned that run this semester")
        return redirect('home')

    context={'course':data2, 't_id': t_id} 

    c_ss = SemesterSession.objects.filter(active = 'Yes').first()

    if request.method == 'POST':
        course = request.POST.get('course_code')
        xx = course.split(",")
        course_id = xx[0]
        dept_id = xx[1]
        marksObj = Result.objects.filter(course_code = course_id, sem_ses = c_ss.ss_id)
        cnt=1
        for i in marksObj:
            i.id = cnt
            cnt = cnt+1

        subject = Subject.objects.get(course_code = course_id)
        subject_name = subject.subject_name
        semester = subject.semester
        subject_dept = subject.dep
        context={'data':marksObj,'course_id':course_id,'subject_name':subject_name,'session':semester,
                'subject_dept':subject_dept, 'student_dept': dept_id
        } 
        return render(request, 'teacher_template/course_result2.html',context)
    return render(request, 'teacher_template/course_result.html',context)

#______________________________Extract Results____________________________________
@login_required(login_url = 'login')
def extract_results(request):
    
    data = Dept.objects.all()

    context={'dept':data, 'level': ['HND1', 'HND2', 'BTECH','M1','M2']} 

    if request.method == 'POST':
        dept_id = request.POST.get('dept')
        level = request.POST.get('level')

        files = []
        f_names = []
        matricules = []
        names = []
        dobs = []
        pobs = []
        fields = ["Matricule", "Names", "Date of Birth", "Place of Birth"]
        f_results = []
        ca_results = []
        exam_results = []

        zip_name = 'Results_' + str(dept_id) + '_' + str(level) + '.zip'


        students = Student.objects.filter(dept = dept_id, level = level).all()

        for r in students:
            matricules.append(r.registration_number)

            names.append(r.name)
            dobs.append(r.dob)
            pobs.append(r.pob)
        
        subjects = Subject.objects.filter(dept = dept_id, level = level, semester = SemesterSession.objects.filter(active = 'Yes').first().semester).all().order_by('subject_name')
        for s in subjects:
            res = []
            ca = []
            exam = []
            res.append(s.subject_name)
            ca.append(s.subject_name)
            exam.append(s.subject_name)
            results = Result.objects.filter(course_code = s.course_code, sem_ses = SemesterSession.objects.filter(active = 'Yes').first()).all()
            if len(results) == 0:
                # create empty results in Result table
                for r in students:
                    new_res = Result(
                        sem_ses = SemesterSession.objects.filter(active = 'Yes').first(),
                        student = r,
                        course_code = s,
                        dept = r.dept,
                        level = r.level,
                    )
                    new_res.save()
            results = Result.objects.filter(course_code = s.course_code, sem_ses = SemesterSession.objects.filter(active = 'Yes').first()).all()
            for r in results:
                if r.resited == "No" :
                    if r.total or r.total != None  :
                        if level == "HND1" or level == "HND2":
                            res.append(r.total)
                        else:
                            res.append(r.total*5)
                    else :
                        res.append(" ")
                    if r.theory_marks or r.theory_marks != None :
                        ca.append(r.theory_marks)
                    else :
                        ca.append(" ")
                    if r.term_test or r.term_test != None :
                        exam.append(r.term_test)
                    else :
                        exam.append(" ")
                else :
                    if level == "HND1" or level == "HND2":
                        res.append(r.total)
                        ca.append(r.theory_marks)
                        exam.append(r.term_test)
                    else:
                        res.append(r.total_resit*5)
                        ca.append(r.theory_marks)
                        exam.append(r.term_test_resit)
                        
            if len(res) > 0 :
                f_results.append(res)
            if len(ca) > 0 :
                ca_results.append(ca)
            if len(exam) > 0 :
                exam_results.append(exam)

        wb_name = 'Combined_Results_' + str(dept_id) + '_' + str(level) + '.xlsx'
        output = io.BytesIO()
        wb = xlsxwriter.Workbook(output)
        ws = wb.add_worksheet('Sheet1')
        
        r1 = 0
        c1 = 0
        
        for f in fields:
            ws.write(r1, c1, f)
            c1+=1
        r2 = 1
        c2 = 0
        
        for (m, n, d, p ) in zip(matricules, names, dobs, pobs):
            ws.write(r2, c2, str(m))
            ws.write(r2,c2 + 1, str(n))
            ws.write(r2,c2 + 2, str(d))
            ws.write(r2,c2 + 3, str(p))
            r2+=1
            
        c3 = 4
        if len(f_results) > 0 :
            for i in range(0, len(f_results)):
                try:
                    r3 = 0
                    for j in range(0, len(matricules)+1):
                        try:
                            ws.write(r3,c3, str(f_results[i][j]))
                            r3+=1
                        except IndexError:
                            continue
                    c3+=1
                except IndexError:
                    continue

        wb.close()
        files.append(output.getvalue())
        f_names.append(wb_name)

        wb_name = 'CA_Results_' + str(dept_id) + '_' + str(level) + '.xlsx'
        output = io.BytesIO()
        wb = xlsxwriter.Workbook(output)
        ws = wb.add_worksheet('Sheet1')
        
        r1 = 0
        c1 = 0
        
        for f in fields:
            ws.write(r1, c1, f)
            c1+=1
        r2 = 1
        c2 = 0
        
        for (m, n, d, p ) in zip(matricules, names, dobs, pobs):
            ws.write(r2, c2, str(m))
            ws.write(r2,c2 + 1, str(n))
            ws.write(r2,c2 + 2, str(d))
            ws.write(r2,c2 + 3, str(p))
            r2+=1
            
        c3 = 4
        if len(ca_results) > 0 :
            for i in range(0, len(ca_results)):
                try:
                    r3 = 0
                    for j in range(0, len(matricules)+1):
                        try:
                            ws.write(r3,c3, ca_results[i][j])
                            r3+=1
                        except IndexError:
                            continue
                    c3+=1
                except IndexError:
                    continue

        wb.close()
        files.append(output.getvalue())
        f_names.append(wb_name)

        wb_name = 'Exam_Results_' + str(dept_id) + '_' + str(level) + '.xlsx'
        output = io.BytesIO()
        wb = xlsxwriter.Workbook(output)
        ws = wb.add_worksheet('Sheet1')
        
        r1 = 0
        c1 = 0
        
        for f in fields:
            ws.write(r1, c1, f)
            c1+=1
        r2 = 1
        c2 = 0
        
        for (m, n, d, p ) in zip(matricules, names, dobs, pobs):
            ws.write(r2, c2, str(m))
            ws.write(r2,c2 + 1, str(n))
            ws.write(r2,c2 + 2, str(d))
            ws.write(r2,c2 + 3, str(p))
            r2+=1
            
        c3 = 4
        if len(exam_results) > 0 :
            for i in range(0, len(exam_results)):
                try:
                    r3 = 0
                    for j in range(0, len(matricules)+1):
                        try:
                            ws.write(r3,c3, exam_results[i][j])
                            r3+=1
                        except IndexError:
                            continue
                    c3+=1
                except IndexError:
                    continue

        wb.close()
        files.append(output.getvalue())
        f_names.append(wb_name)


        if len(files) > 1:
            import zipfile
            output = io.BytesIO()
            with zipfile.ZipFile(output, 'w') as zf:
                for n,p in zip(f_names, files):
                    zf.writestr(n,p)
                zf.close()

            response = HttpResponse(
            output.getvalue(),
                content_type="application/x-zip-compressed",
            )
            response["Content-Disposition"] = "attachment; filename=" + zip_name
                    
                    # rendering the template
            return response


    return render(request,'admin_template/extract_results.html',context)

#__________________________________________________EXTRACT RESULTS END_____________________________________________________

@login_required(login_url = 'login')
@allowed_users(allowed_roles=['admin'])
def generate_codes(request):

    import random

    sem = SemesterSession.objects.filter(active = 'Yes').first()

    data = Subject.objects.filter(semester = sem.semester)

    context = {'course': data}

    if request.method == 'POST':
        course_code = request.POST.get('course_code').split(",")[0]
        

        studs = RegisterTable.objects.filter(subject = course_code, sem_ses = sem).all()
        my_list = list(range(1,int(studs.count()) + 1))
        success = False
        for stud in studs:
            try :
                code = random.choice(my_list)
                my_list.remove(code)

                exam_code = ExamCode(
                    sem_ses = sem,
                    subject = Subject.objects.filter(course_code = course_code).first(),
                    student = stud.student,
                    code = code
                )

                exam_code.save()
                success = True
            except IntegrityError:
                messages.success(request,"Student %s already has a code for this course. "%(stud.student.name))
                success = True
                continue
        if success:    
            messages.success(request,"Successfully Generated Codes For Course ")

    return render(request,'admin_template/generate_codes.html',context)

@login_required(login_url = 'login')
@allowed_users(allowed_roles=['admin'])
def download_codes(request):

    sem = SemesterSession.objects.filter(active = 'Yes').first()

    data = Subject.objects.filter(semester = sem.semester)

    context = {'course': data}

    if request.method == 'POST':
        course_code = request.POST.get('course_code').split(",")[0]

        course_name = Subject.objects.filter(course_code = course_code).first().subject_name

        matricules = []
        names = []
        codes = []
        fields = ["Matricule", "Names", "Code"]

        register1 = RegisterTable.objects.filter(subject_id = course_code, sem_ses = sem).all()
        for r in register1:
            matricules.append(r.student)

            n = Student.objects.filter(registration_number = r.student).first().name
            names.append(n)

            c = ExamCode.objects.filter(student = r.student, sem_ses = sem, subject = r.subject).first().code
            codes.append(c)

        wb_name = 'Codes_' + str(course_name) + '_' + str(sem.ss_id) +'.xlsx'

        output = io.BytesIO()
        wb = xlsxwriter.Workbook(output)
        ws = wb.add_worksheet('Sheet1')
        
        r1 = 0
        c1 = 0

        for f in fields:
            ws.write(r1, c1, f)
            c1+=1
        r2 = 1
        c2 = 0

        for (m, n, c) in zip(matricules, names, codes):
            ws.write(r2, c2, str(m))
            ws.write(r2,c2 + 1, str(n))
            ws.write(r2,c2 + 2, str(c))
            r2+=1

        wb.close()
        
        output.seek(0)

        response = HttpResponse(
            output,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = "attachment; filename=%s" % wb_name

        return response
            

    return render(request,'admin_template/download_codes.html',context)


def delete_result(request):
    t_id = str(request.user.teacher.teacher_id)
    data = AssignedTeacher2.objects.filter(teacher_id = t_id)
    
    context = {'data':data}
    if request.method == 'POST':
        course_id = request.POST.get('course_code')
        xx = course_id.split(",")
        
        return redirect(reverse('delete_result2', kwargs= {"dept": xx[1], "course_id": xx[0]}))

    return render(request,'teacher_template/delete_result.html',context)


def delete_result2(request, dept, course_id):
    data = Result.objects.filter(dept = dept, course_code = course_id)

    if request.method == 'POST':
        course_id = request.POST.get('course_code')
        xx = course_id.split(",")

        Result.objects.filter(course_code = xx[1], student_id = xx[0]).delete()

        messages.success(request,"Successfully Deleted %s student's %s Course Result"%(xx[0],xx[1]))
        return redirect('home')
    
    return render(request,'teacher_template/delete_result.html')

def delete_student(request):
    data = Student.objects.all()

    if request.method == 'POST':
        regi = request.POST.get('course_code')
    
        uid = Student.objects.get(registration_number = regi).user_id
        Student.objects.filter(registration_number = regi).delete()
        User.objects.filter(id = uid).delete()

        messages.success(request,"Successfully Deleted %s student"%(regi))
        return redirect('home')

    context ={
        'data':data


    }
    return render(request,'admin_template/remove_student.html',context)

def remove_teacher(request):
    data = Teacher.objects.all()

    if request.method == 'POST':
        regi = request.POST.get('course_code')
    
        uid = Teacher.objects.get(teacher_id = regi).user_id
        Teacher.objects.filter(teacher_id = regi).delete()
        User.objects.filter(id = uid).delete()

        messages.success(request,"Successfully Deleted %s Teacher"%(regi))
        return redirect('home')

    context ={
        'data':data


    }
    return render(request,'admin_template/remove_teacher.html',context)

class GeneratePdf2(View):
    course_id = None
    dept_id = None
    def get(self, request, *args, **kwargs):
        self.course_id = self.kwargs.get('course_id', None)
        self.dept_id = self.kwargs.get('dept_id',None)

        cid = str(self.course_id)
        did = str(self.dept_id)
        marksObj = Result.objects.filter(course_code = cid, sem_ses = SemesterSession.objects.filter(active= 'Yes').first())
        cnt=1
        for i in marksObj:
            i.id = cnt
            cnt = cnt+1
            i.cgname = cal_grade(i.total) 

    
        dept_name = Dept.objects.get(dept_id = did).name
        subject_name = Subject.objects.get(course_code = cid).subject_name 
        credit = Subject.objects.get(course_code = cid).credit 
        session  = Subject.objects.get(course_code = cid).semester

        module_dir = os.path.dirname(__file__)  # get current directory
        file_path1 = os.path.join(module_dir, 'templates/teacher_template/generate_result_pdf_temp.html')
        file_path2 = os.path.join(module_dir, 'templates/teacher_template/generate_result_pdf.html')

        pwd = os.path.dirname(__file__)
        open(file_path1, "w").write(render_to_string(file_path2, {'data': marksObj, 'course_number': cid,'dept_name': dept_name,'subject_name': subject_name,'credit':credit ,'session':session}))

        # Converting the HTML template into a PDF file
        pdf = html_to_pdf(file_path1)

        # rendering the template
        return HttpResponse(pdf, content_type='application/pdf')

from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.forms import PasswordChangeForm

@login_required(login_url = 'login')
def change_pwd(request):
    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)  # Important!
            messages.success(request, 'Your password was successfully updated!')
            return redirect('logout')
        else:
            messages.error(request, 'Please correct the error below.')
    else:
        form = PasswordChangeForm(request.user)
    return render(request, 'main/change_password.html', {
        'form': form
    })

@login_required(login_url = 'login')
def stud_update_info(request):
    user = request.user.student.user
    u =  get_object_or_404(User, username = user.username )
    stud = request.user.student
    student = get_object_or_404(Student, registration_number = stud.registration_number )
    form = StudentUpdateForm(request.POST or None, instance = student)

    c_ss = SemesterSession.objects.filter(active = "Yes").first()
    context = {'form': form, 'aca': c_ss.session, 'email': user.email}

    if request.method == 'POST':
        email = request.POST.get('email')
        if form.is_valid() :
            u.username = email
            u.email = email
            u.save()
            form.save()
            student.save()

            messages.info(request,"Info Successfully Updated! Your username is now the entered email.")
            return redirect('home')
    
    return render(request, 'student_template/update_info.html', context)

@login_required(login_url="login")
@allowed_users(allowed_roles=['admin'])
def sem_update(request):
    c_ss = SemesterSession.objects.filter(active = 'Yes').first()
    sem = get_object_or_404(SemesterSession,ss_id = c_ss.ss_id)
    form = SSUpdateForm(request.POST or None, instance = sem)

    context = {"form": form}
    if request.method == 'POST':
        if form.is_valid():
            form.save()
            sem.save()

            messages.info(request,"Info Successfully Updated!")
            return redirect('home')
    
    return render(request, 'admin_template/update_sem_info.html', context)
